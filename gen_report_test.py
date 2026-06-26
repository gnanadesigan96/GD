"""
gen_report_test.py
TEST SCRIPT — extends gen_report_live.py with an Alert Summary widget.
Once approved, merge the alert logic into gen_report_live.py.

Extra widget added after ticket detail section:
  - Alert breakdown by Environment × Alert Type
  - Today count | Yesterday count | Last 7-day count
  - Next-day forecast (7-day average)

Alerts are identified by email = notify-sre-ops@corestack.io
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "azure-function"))

import logging
import re
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timezone, timedelta

import io
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from report_generator import parse_ticket, generate_html, generate_excel
from sharepoint_client import upload_file

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

# ── Credentials ────────────────────────────────────────────────────────────────
ZOHO_CLIENT_ID     = os.environ.get("ZOHO_CLIENT_ID",     "1000.LXE6HGZAW4FWRED50ZUZ42CHUFHVEO")
ZOHO_CLIENT_SECRET = os.environ.get("ZOHO_CLIENT_SECRET", "7d84eb43d93d42648ad05636b2b7310652361722e9")
ZOHO_REFRESH_TOKEN = os.environ.get("ZOHO_REFRESH_TOKEN", "1000.ca586bbc38413c0661c0e67e78378449.72ac1952b7eb013374cc87e8544475a4")

SHAREPOINT_TENANT_ID     = os.environ.get("SHAREPOINT_TENANT_ID",     "")
SHAREPOINT_CLIENT_ID     = os.environ.get("SHAREPOINT_CLIENT_ID",     "abb2a8fa-4603-4aff-80b2-bf614beb173b")
SHAREPOINT_CLIENT_SECRET = os.environ.get("SHAREPOINT_CLIENT_SECRET", "")
SHAREPOINT_SITE_URL      = os.environ.get("SHAREPOINT_SITE_URL",      "cloudenablersinc.sharepoint.com/sites/SupportTeam")
SHAREPOINT_HTML_FOLDER   = os.environ.get("SHAREPOINT_HTML_FOLDER",   "General/Daily-Incident-Report/Template")
SHAREPOINT_EXCEL_FOLDER  = os.environ.get("SHAREPOINT_EXCEL_FOLDER",  "General/Daily-Incident-Report/Excel")

try:
    from credentials_zoho import (        # type: ignore
        SHAREPOINT_TENANT_ID     as _T,
        SHAREPOINT_CLIENT_SECRET as _S,
    )
    if _T: SHAREPOINT_TENANT_ID     = _T
    if _S: SHAREPOINT_CLIENT_SECRET = _S
except ImportError:
    pass

os.environ["SHAREPOINT_TENANT_ID"]     = SHAREPOINT_TENANT_ID
os.environ["SHAREPOINT_CLIENT_ID"]     = SHAREPOINT_CLIENT_ID
os.environ["SHAREPOINT_CLIENT_SECRET"] = SHAREPOINT_CLIENT_SECRET
os.environ["SHAREPOINT_SITE_URL"]      = SHAREPOINT_SITE_URL
os.environ["SHAREPOINT_HTML_FOLDER"]   = SHAREPOINT_HTML_FOLDER
os.environ["SHAREPOINT_EXCEL_FOLDER"]  = SHAREPOINT_EXCEL_FOLDER

ZOHO_ORG_ID        = "60019389025"
ZOHO_DEPT_ID       = "100599000000010772"
PENTAGON_ACCOUNT   = "100599000037212179"
ALERT_EMAIL        = "notify-sre-ops@corestack.io"

ZOHO_TOKEN_URL = "https://accounts.zoho.in/oauth/v2/token"
ZOHO_API_BASE  = "https://desk.zoho.in/api/v1"
IST = timedelta(hours=5, minutes=30)


# ── Zoho helpers ───────────────────────────────────────────────────────────────
def get_token() -> str:
    resp = requests.post(ZOHO_TOKEN_URL, params={
        "refresh_token": ZOHO_REFRESH_TOKEN,
        "client_id":     ZOHO_CLIENT_ID,
        "client_secret": ZOHO_CLIENT_SECRET,
        "grant_type":    "refresh_token",
    }, timeout=30)
    resp.raise_for_status()
    return resp.json()["access_token"]


def hdrs(token: str) -> dict:
    return {"Authorization": f"Zoho-oauthtoken {token}", "orgId": ZOHO_ORG_ID}


def fetch_by_status(status: str, token: str) -> list[dict]:
    tickets, from_ = [], 0
    while True:
        resp = requests.get(f"{ZOHO_API_BASE}/tickets", headers=hdrs(token), params={
            "departmentId": ZOHO_DEPT_ID, "status": status,
            "sortBy": "createdTime", "limit": 50, "from": from_,
        }, timeout=30)
        resp.raise_for_status()
        data = resp.json().get("data", [])
        if not data: break
        tickets.extend(data)
        if len(data) < 50: break
        from_ += 50
    return tickets


def fetch_detail(ticket_id: str, token: str) -> dict:
    resp = requests.get(
        f"{ZOHO_API_BASE}/tickets/{ticket_id}",
        headers=hdrs(token),
        params={"include": "contacts,assignee"},
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


def fetch_all_recent_alerts(token: str, days_back: int = 8) -> list[dict]:
    """
    Fetch all SRE alert tickets from the last `days_back` days.
    Zoho doesn't support createdTimeRange filtering, so we fetch all statuses
    sorted by createdTime desc and stop once tickets are older than the cutoff.
    """
    cutoff_ist = datetime.now(timezone.utc) + IST - timedelta(days=days_back)
    cutoff_str = cutoff_ist.strftime("%Y-%m-%dT%H:%M:%S")

    results = []
    statuses = ["Open", "In Progress", "On Hold", "Awaiting Resolution Confirmation", "Closed"]
    for status in statuses:
        from_ = 0
        while True:
            resp = requests.get(f"{ZOHO_API_BASE}/tickets", headers=hdrs(token), params={
                "departmentId": ZOHO_DEPT_ID,
                "status":       status,
                "sortBy":       "createdTime",
                "limit":        100,
                "from":         from_,
            }, timeout=30)
            resp.raise_for_status()
            data = resp.json().get("data", [])
            if not data:
                break
            # Filter to alert email and within cutoff window
            for t in data:
                email = (t.get("email") or t.get("contactEmail") or "").lower()
                ct = (t.get("createdTime") or "")[:19]
                if email == ALERT_EMAIL and ct >= cutoff_str:
                    results.append(t)
            if len(data) < 100:
                break
            from_ += 100
    return results


def _ticket_ist_date(t: dict) -> date:
    """Parse ticket createdTime (UTC) and return the IST calendar date."""
    ct = t.get("createdTime") or ""
    if not ct:
        return date.min
    # Format: "2026-06-26T10:30:00.000Z"
    try:
        dt_utc = datetime.strptime(ct[:19], "%Y-%m-%dT%H:%M:%S").replace(tzinfo=timezone.utc)
        return (dt_utc + IST).date()
    except ValueError:
        return date.min


# ── Alert parsing ──────────────────────────────────────────────────────────────
_ENV_PATTERNS = [
    # Specific app-tier envs — must come before broader patterns
    (r"prd-us3-app|prd-us3-web",                                           "prd-us3-app"),
    (r"prd-us-app",                                                        "prd-us-app"),
    (r"prd-eu-app",                                                        "prd-eu-app"),
    (r"prd-in-app",                                                        "prd-in-app"),
    (r"prd-uae-app",                                                       "prd-uae-app"),
    # Broader env buckets
    (r"prd-us3|us3|useast|us4|useast-web|us-east|blackstone",             "USEast"),
    (r"prd-us-web|prd-us\b|prod-us\b|prod-us-",                          "ProdUS"),
    (r"prd-eu|prod-eu|eu\.corestack",                                     "ProdEU"),
    (r"prd-uae|prod-uae|uae",                                             "ProdUAE"),
    (r"msprod|ms[-\s]?prod",                                              "MSProd"),
    (r"kyndryl",                                                          "Kyndryl"),
    (r"prodin|prd-in|prod-in",                                            "ProdIN"),
]

_ALERT_PATTERNS = [
    (r"cpu usage|cpu utilisation|cpu utilization|cpu",                    "CPU"),
    (r"pod.*memory|memory.*pod",                                          "Pod Memory"),
    (r"memory usage|memory utilisation|memory utilization|memory",        "Memory"),
    (r"oomkilled|oom",                                                    "OOM"),
    (r"vm availability|availability is below",                            "VM Availability"),
    (r"missing service|service.*down",                                    "Service Down"),
    (r"url.*down|is down|unreachable",                                    "URL Down"),
    (r"disk|storage",                                                     "Disk"),
    (r"pod.*restart|restartcount",                                        "Pod Restart"),
    (r"db alert|database",                                                "Database"),
    (r"graphdb|graph db|graph database",                                  "GraphDB"),
    (r"node",                                                             "Node"),
    (r"network",                                                          "Network"),
]


def parse_alert(subject: str) -> tuple[str, str]:
    """Return (environment, alert_type) from subject line."""
    s = subject.lower()
    env = "Other"
    for pattern, label in _ENV_PATTERNS:
        if re.search(pattern, s):
            env = label
            break
    alert_type = "Other"
    for pattern, label in _ALERT_PATTERNS:
        if re.search(pattern, s):
            alert_type = label
            break
    return env, alert_type


# ── Alert HTML widget ─────────────────────────────────────────────────────────
def build_alert_widget(today_alerts: list, yesterday_alerts: list, week_alerts: list) -> str:
    """Build the HTML alert summary widget — one mini-table per environment."""

    def count_by(alerts):
        counts = defaultdict(lambda: defaultdict(int))
        for t in alerts:
            env, atype = parse_alert(t.get("subject") or "")
            counts[env][atype] += 1
        return counts

    today_c     = count_by(today_alerts)
    yesterday_c = count_by(yesterday_alerts)
    week_c      = count_by(week_alerts)

    # Log "Other" subjects so we can improve patterns
    for t in today_alerts + yesterday_alerts + week_alerts:
        env, atype = parse_alert(t.get("subject") or "")
        if env == "Other" or atype == "Other":
            logging.info("Alert [%s/%s]: %s", env, atype, t.get("subject", ""))

    def _sort_key(x):
        return (x == "Other", x)

    all_envs  = sorted(set(list(today_c) + list(yesterday_c) + list(week_c)), key=_sort_key)
    all_types = sorted(set(
        tp for c in [today_c, yesterday_c, week_c] for env in c for tp in c[env]
    ), key=_sort_key)

    n_today     = len(today_alerts)
    n_yesterday = len(yesterday_alerts)

    if not all_envs:
        return (
            '<tr><td style="padding-bottom:14px;">'
            '<table width="100%" cellpadding="6" cellspacing="0" border="0" '
            'style="background:#fff;border:1px solid #E4E8EF;">'
            '<tr><td style="background:#F8FAFC;border-bottom:1px solid #E4E8EF;padding:8px 14px;">'
            '<span style="font-size:12px;font-weight:700;color:#0F172A;">&#128680; SRE Alert Summary</span>'
            '</td></tr>'
            '<tr><td style="padding:12px;font-size:11px;color:#64748B;">No alerts recorded today.</td></tr>'
            '</table></td></tr>\n'
        )

    # Overall trend badge
    if n_today > n_yesterday:
        trend_badge = (f' <span style="font-size:10px;font-weight:700;color:#B91C1C;">'
                       f'&#9650; +{n_today - n_yesterday}</span>')
    elif n_today < n_yesterday:
        trend_badge = (f' <span style="font-size:10px;font-weight:700;color:#16A34A;">'
                       f'&#9660; {n_today - n_yesterday}</span>')
    else:
        trend_badge = ""

    # ── Per-environment mini-tables laid out in a 3-column grid ──────────────
    # Env header accent colours (cycles for variety)
    _ENV_COLORS = ["#1D4ED8", "#0369A1", "#6D28D9", "#0F766E", "#B45309", "#BE185D", "#374151"]

    def _mini_table(env: str, color: str) -> str:
        rows = ""
        for j, atype in enumerate(all_types):
            td = today_c.get(env, {}).get(atype, 0)
            yd = yesterday_c.get(env, {}).get(atype, 0)
            wd = week_c.get(env, {}).get(atype, 0)
            if td == 0 and yd == 0 and wd == 0:
                continue
            if td == 0 and yd == 0:
                today_cell = f'<span style="font-size:12px;color:#CBD5E1;">&ndash;</span>'
            elif td > yd:
                today_cell = (
                    f'<span style="font-size:13px;font-weight:800;color:#DC2626;">{td}</span>'
                    f'&thinsp;<span style="font-size:9px;color:#DC2626;vertical-align:middle;">&#9650;</span>'
                )
            elif td < yd:
                today_cell = (
                    f'<span style="font-size:13px;font-weight:800;color:#16A34A;">{td}</span>'
                    f'&thinsp;<span style="font-size:9px;color:#16A34A;vertical-align:middle;">&#9660;</span>'
                )
            elif td == 0:
                # today 0, yd also 0 handled above; here yd>0 means it dropped to 0
                today_cell = (
                    f'<span style="font-size:13px;font-weight:800;color:#16A34A;">0</span>'
                    f'&thinsp;<span style="font-size:9px;color:#16A34A;vertical-align:middle;">&#9660;</span>'
                )
            else:
                # equal and both non-zero — steady
                today_cell = (
                    f'<span style="font-size:13px;font-weight:800;color:#1E293B;">{td}</span>'
                    f'&thinsp;<span style="font-size:9px;color:#94A3B8;vertical-align:middle;">=</span>'
                )
            bg = "#F8FAFC" if j % 2 == 0 else "#FFFFFF"
            rows += (
                f'<tr style="background:{bg};">'
                f'<td style="padding:5px 10px;font-size:11px;color:#374151;'
                f'border-bottom:1px solid #F1F5F9;max-width:130px;">{atype}</td>'
                f'<td style="padding:5px 10px;text-align:center;border-bottom:1px solid #F1F5F9;">{today_cell}</td>'
                f'<td style="padding:5px 10px;text-align:center;font-size:11px;color:#6B7280;'
                f'border-bottom:1px solid #F1F5F9;">{yd if yd else "&ndash;"}</td>'
                f'<td style="padding:5px 10px;text-align:center;font-size:11px;color:#94A3B8;'
                f'border-bottom:1px solid #F1F5F9;">{wd if wd else "&ndash;"}</td>'
                f'</tr>'
            )
        if not rows:
            return ""
        env_today = sum(today_c.get(env, {}).values())
        env_yest  = sum(yesterday_c.get(env, {}).values())
        if env_today > env_yest:
            env_dot = '<span style="font-size:9px;color:#DC2626;font-weight:700;"> &#9650;</span>'
        elif env_today < env_yest:
            env_dot = '<span style="font-size:9px;color:#16A34A;font-weight:700;"> &#9660;</span>'
        else:
            env_dot = ""
        return (
            '<table width="100%" cellpadding="0" cellspacing="0" border="0" '
            f'style="border:1px solid #E2E8F0;border-top:3px solid {color};border-radius:0 0 4px 4px;">'
            # env name header
            f'<tr><td colspan="4" style="padding:8px 10px 6px 10px;background:#FAFBFC;">'
            f'<span style="font-size:12px;font-weight:700;color:{color};">{env}</span>'
            f'<span style="font-size:11px;color:#64748B;"> — {env_today} today{env_dot}</span>'
            f'</td></tr>'
            # column headers
            '<tr style="background:#F1F5F9;">'
            '<th style="padding:4px 10px;font-size:9px;font-weight:600;color:#94A3B8;'
            'text-transform:uppercase;text-align:left;border-bottom:1px solid #E2E8F0;">Type</th>'
            '<th style="padding:4px 10px;font-size:9px;font-weight:600;color:#94A3B8;'
            'text-transform:uppercase;text-align:center;border-bottom:1px solid #E2E8F0;">Today</th>'
            '<th style="padding:4px 10px;font-size:9px;font-weight:600;color:#94A3B8;'
            'text-transform:uppercase;text-align:center;border-bottom:1px solid #E2E8F0;">Yest.</th>'
            '<th style="padding:4px 10px;font-size:9px;font-weight:600;color:#94A3B8;'
            'text-transform:uppercase;text-align:center;border-bottom:1px solid #E2E8F0;">7d</th>'
            '</tr>'
            + rows
            + '</table>'
        )

    # Build list of non-empty mini-tables
    mini_tables = []
    for idx, env in enumerate(all_envs):
        color = _ENV_COLORS[idx % len(_ENV_COLORS)]
        mt = _mini_table(env, color)
        if mt:
            mini_tables.append(mt)

    # 3-column grid
    grid_rows = ""
    for i in range(0, len(mini_tables), 3):
        chunk = mini_tables[i:i+3]
        while len(chunk) < 3:
            chunk.append("")
        grid_rows += '<tr valign="top">'
        for mt in chunk:
            grid_rows += (
                '<td style="width:33%;padding:0 8px 16px 0;vertical-align:top;">'
                + mt
                + '</td>'
            )
        grid_rows += '</tr>'

    # Summary chips for header
    def _chip(label, val, bg, fg):
        return (
            f'<span style="display:inline-block;background:{bg};color:{fg};'
            f'font-size:10px;font-weight:600;padding:3px 10px;border-radius:12px;margin-left:6px;">'
            f'{label}: {val}</span>'
        )

    widget = (
        '<tr><td style="padding-bottom:20px;">'
        '<table width="100%" cellpadding="0" cellspacing="0" border="0" '
        'style="background:#fff;border:1px solid #E2E8F0;border-radius:6px;overflow:hidden;">'

        # Header
        '<tr><td style="background:linear-gradient(135deg,#1E293B 0%,#334155 100%);padding:11px 16px;">'
        '<table width="100%" cellpadding="0" cellspacing="0" border="0"><tr>'
        '<td style="font-size:13px;font-weight:700;color:#F8FAFC;letter-spacing:.3px;">'
        '&#128680;&nbsp; SRE Alert Summary</td>'
        f'<td align="right">'
        + _chip("Today",     n_today,           "#FEE2E2", "#991B1B")
        + (f'<span style="display:inline-block;font-size:10px;font-weight:700;'
           f'color:{"#DC2626" if n_today > n_yesterday else "#16A34A"};margin-left:4px;">'
           f'{"&#9650; +" if n_today > n_yesterday else "&#9660; "}{abs(n_today - n_yesterday)}</span>'
           if n_today != n_yesterday else "")
        + _chip("Yesterday",  n_yesterday,        "#FEF3C7", "#92400E")
        + _chip("Last 7 days", len(week_alerts),  "#DBEAFE", "#1E40AF")
        + f'</td>'
        '</tr></table></td></tr>'

        # Grid
        '<tr><td style="padding:14px 12px 2px 12px;">'
        '<table width="100%" cellpadding="0" cellspacing="0" border="0">'
        + grid_rows +
        '</table></td></tr>'
        '</table></td></tr>\n'
    )
    return widget


def build_ticket_stats_footer(
    tickets: list,
    today: date,
    cur_week_raw: list,
    last_week_raw: list,
    month_raw: list,
) -> str:
    """Footer showing ticket volume: this week vs last week vs this month."""
    cur_w  = len(cur_week_raw)
    last_w = len(last_week_raw)
    month  = len(month_raw)

    if cur_w > last_w:
        w_trend = (f'<span style="font-size:10px;color:#DC2626;font-weight:700;">'
                   f' &#9650; +{cur_w - last_w}</span>')
    elif cur_w < last_w:
        w_trend = (f'<span style="font-size:10px;color:#16A34A;font-weight:700;">'
                   f' &#9660; {cur_w - last_w}</span>')
    else:
        w_trend = ""

    def _col(icon, label, value, extra, bg, accent):
        return (
            f'<td style="width:33%;padding:0 8px;vertical-align:top;">'
            f'<table width="100%" cellpadding="0" cellspacing="0" border="0" '
            f'style="border:1px solid #E2E8F0;border-top:3px solid {accent};border-radius:0 0 4px 4px;">'
            f'<tr><td style="padding:12px 14px;background:#FAFBFC;">'
            f'<div style="font-size:10px;color:#94A3B8;text-transform:uppercase;'
            f'letter-spacing:.6px;margin-bottom:6px;">{icon} {label}</div>'
            f'<div style="font-size:26px;font-weight:800;color:#0F172A;line-height:1;">'
            f'{value}{extra}</div>'
            f'</td></tr>'
            f'</table></td>'
        )

    return (
        '<tr><td style="padding-bottom:20px;">'
        '<table width="100%" cellpadding="0" cellspacing="0" border="0" '
        'style="background:#fff;border:1px solid #E2E8F0;border-radius:6px;">'
        '<tr><td style="background:#F8FAFC;border-bottom:1px solid #E2E8F0;padding:9px 16px;">'
        '<span style="font-size:12px;font-weight:700;color:#0F172A;">&#128202; Customer Incident Volume</span>'
        f'<span style="font-size:10px;color:#94A3B8;margin-left:8px;">{today.strftime("%B %Y")} &mdash; all customer incidents</span>'
        '</td></tr>'
        '<tr><td style="padding:12px 8px;">'
        '<table width="100%" cellpadding="0" cellspacing="0" border="0"><tr>'
        + _col("&#128197;", "This Week",        cur_w,  w_trend,  "#EFF6FF", "#2563EB")
        + _col("&#128336;", "Last Week",        last_w, "",       "#F8FAFC", "#64748B")
        + _col("&#128218;", f"{today.strftime('%B')} Total", month, "", "#FFF7ED", "#D97706")
        + '</tr></table>'
        '</td></tr>'
        '</table></td></tr>\n'
    )


def generate_html_with_alerts(tickets, today, excel_url, alert_widget) -> str:
    """Generate the standard HTML and inject the alert widget after the ticket section."""
    html = generate_html(tickets, today, excel_url=excel_url)
    inject_before = '<tr><td><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#F0F9FF'
    html = html.replace(inject_before, alert_widget + inject_before, 1)
    return html


def add_alerts_sheet(excel_bytes: bytes, today_alerts: list, today: date) -> bytes:
    """Append an 'Alerts' sheet with today's alert ticket details to the workbook bytes."""
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb.create_sheet("Alerts")

    hdr_fill = PatternFill("solid", fgColor="1E293B")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    cols = ["Ticket #", "Subject", "Environment", "Alert Type", "Status", "Created Time (IST)"]
    widths = [12, 60, 16, 18, 14, 22]
    for i, (col, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=1, column=i, value=col)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    for r, t in enumerate(today_alerts, 2):
        env, atype = parse_alert(t.get("subject") or "")
        ct = t.get("createdTime") or ""
        ist_str = ""
        if ct:
            try:
                dt_utc = datetime.strptime(ct[:19], "%Y-%m-%dT%H:%M:%S").replace(tzinfo=timezone.utc)
                ist_str = (dt_utc + IST).strftime("%Y-%m-%d %H:%M IST")
            except ValueError:
                ist_str = ct
        row_data = [
            t.get("ticketNumber") or t.get("id") or "",
            t.get("subject") or "",
            env,
            atype,
            t.get("status") or "",
            ist_str,
        ]
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c_idx, value=val)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def fetch_tickets_in_range(token: str, from_date: date, to_date: date) -> list[dict]:
    """Fetch all non-noise tickets created between from_date and to_date (IST) across all statuses."""
    statuses = ["Open", "In Progress", "On Hold", "Awaiting Resolution Confirmation", "Closed"]
    cutoff_from = datetime(from_date.year, from_date.month, from_date.day, 0, 0, 0, tzinfo=timezone.utc) - IST
    cutoff_to   = datetime(to_date.year,   to_date.month,   to_date.day,   23, 59, 59, tzinfo=timezone.utc) - IST
    from_str = cutoff_from.strftime("%Y-%m-%dT%H:%M:%S")
    to_str   = cutoff_to.strftime("%Y-%m-%dT%H:%M:%S")

    results = []
    for status in statuses:
        from_ = 0
        while True:
            resp = requests.get(f"{ZOHO_API_BASE}/tickets", headers=hdrs(token), params={
                "departmentId": ZOHO_DEPT_ID, "status": status,
                "sortBy": "createdTime", "limit": 100, "from": from_,
            }, timeout=30)
            resp.raise_for_status()
            data = resp.json().get("data", [])
            if not data:
                break
            for t in data:
                ct = (t.get("createdTime") or "")[:19]
                if from_str <= ct <= to_str:
                    results.append(t)
            if len(data) < 100:
                break
            from_ += 100
    return results


def is_pentagon(t: dict) -> bool:
    acc = (t.get("account") or {})
    acc_id = acc.get("id") or acc.get("accountId") or t.get("accountId") or ""
    if str(acc_id) == PENTAGON_ACCOUNT: return True
    subj  = (t.get("subject") or "").lower()
    email = (t.get("email") or t.get("contactEmail") or "").lower()
    if "clouddesk@pentagon" in email: return True
    if "[inc-52" in subj and ("request resolved" in subj or "request received" in subj): return True
    return False


def is_alert(t: dict) -> bool:
    email = (t.get("email") or t.get("contactEmail") or "").lower()
    return ALERT_EMAIL in email


def main():
    today_ist = datetime.now(timezone.utc) + IST
    today     = today_ist.date()
    logging.info("Generating TEST report for %s", today)

    token = get_token()
    logging.info("Token obtained.")

    # ── Fetch active incident tickets ─────────────────────────────────────────
    statuses = ["Open", "In Progress", "On Hold", "Awaiting Resolution Confirmation"]
    raw = []
    for s in statuses:
        batch = fetch_by_status(s, token)
        logging.info("  %s: %d tickets", s, len(batch))
        raw.extend(batch)

    raw = [t for t in raw if not is_pentagon(t) and not is_alert(t)]
    logging.info("After filters: %d tickets", len(raw))

    def enrich(t: dict) -> dict:
        tid = t.get("id") or t.get("ticketId") or ""
        if not tid: return t
        try:
            detail = fetch_detail(str(tid), token)
            t["cf"] = detail.get("cf") or {}
            t["customFields"] = detail.get("customFields") or {}
            for f in ("contacts", "contact", "assignee", "account"):
                if detail.get(f): t[f] = detail[f]
            if detail.get("resolution") and not t.get("resolution"):
                t["resolution"] = detail["resolution"]
            if detail.get("ticketType"): t["ticketType"] = detail["ticketType"]
        except Exception as e:
            logging.warning("Detail fetch failed for %s: %s", tid, e)
        return t

    logging.info("Enriching tickets…")
    with ThreadPoolExecutor(max_workers=10) as ex:
        futures = {ex.submit(enrich, t): t for t in raw}
        enriched = [f.result() for f in as_completed(futures)]
    logging.info("Enrichment done.")

    enriched = [t for t in enriched
                if ((t.get("cf") or {}).get("cf_request_type") or "").lower() in ("incident request", "")]
    logging.info("After incident-only filter: %d tickets", len(enriched))

    tickets = [parse_ticket(t, today) for t in enriched]

    # ── Fetch alert tickets for today / yesterday / last 7 days ───────────────
    logging.info("Fetching SRE alert tickets…")
    all_alerts = fetch_all_recent_alerts(token, days_back=8)
    logging.info("Total recent alert tickets fetched: %d", len(all_alerts))

    yesterday = today - timedelta(days=1)
    today_alerts     = [t for t in all_alerts if _ticket_ist_date(t) == today]
    yesterday_alerts = [t for t in all_alerts if _ticket_ist_date(t) == yesterday]
    week_alerts      = [t for t in all_alerts
                        if yesterday >= _ticket_ist_date(t) >= today - timedelta(days=7)]

    logging.info("Alerts — today: %d  yesterday: %d  last 7 days: %d",
                 len(today_alerts), len(yesterday_alerts), len(week_alerts))

    alert_widget = build_alert_widget(today_alerts, yesterday_alerts, week_alerts)

    # ── Generate files ────────────────────────────────────────────────────────
    date_tag   = today.strftime("%Y-%m-%d")
    html_path  = f"CS_Daily_Incident_Report_{date_tag}_TEST.html"
    excel_path = f"CS_Daily_Incident_Report_{date_tag}_TEST.xlsx"

    excel_bytes = generate_excel(tickets, today)
    try:
        excel_bytes = add_alerts_sheet(excel_bytes, today_alerts, today)
        logging.info("Alerts sheet added (%d rows)", len(today_alerts))
    except Exception as e:
        logging.error("Failed to add Alerts sheet: %s", e)
    with open(excel_path, "wb") as f:
        f.write(excel_bytes)
    logging.info("Excel written: %s", excel_path)

    sp_html_folder  = os.environ["SHAREPOINT_HTML_FOLDER"]
    sp_excel_folder = os.environ["SHAREPOINT_EXCEL_FOLDER"]
    excel_sp_url = ""
    try:
        logging.info("Uploading Excel to SharePoint…")
        excel_sp_url = upload_file(excel_path, sp_excel_folder, excel_path)
        logging.info("Excel uploaded: %s", excel_sp_url)
    except Exception as e:
        logging.error("SharePoint Excel upload failed: %s", e)

    html_content = generate_html_with_alerts(tickets, today, excel_sp_url, alert_widget)
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)
    logging.info("HTML written: %s", html_path)

    try:
        logging.info("Uploading HTML to SharePoint…")
        html_url = upload_file(html_path, sp_html_folder, html_path)
        logging.info("HTML uploaded: %s", html_url)
    except Exception as e:
        logging.error("SharePoint HTML upload failed: %s", e)
        logging.info("Files saved locally.")

    logging.info("Done. %d tickets | Today alerts: %d", len(tickets), len(today_alerts))


if __name__ == "__main__":
    main()
