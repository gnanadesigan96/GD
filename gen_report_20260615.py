#!/usr/bin/env python3
"""Generate CS_Daily_Incident_Report_20260615.html and .xlsx"""

import os

# ── DATA ────────────────────────────────────────────────────────────────────

TICKETS = [
    # fmt: ticket, subject, priority, account, band, ado, raised_by, assignee, last_update, status, age_days, hold_reason
    # IN PROGRESS
    {"id":"211192","subj":"FinOps Email Notification Issue (Clone)","pri":"P3","acct":"Tata Communications","band":"Gold","ado":"125029","raised":"Sunilkumar S","team":"PremKumar B","upd":"Jun 9","status":"In Progress","age":27,"reason":""},
    {"id":"217515","subj":"FW: Cost Mismatch in recommendation","pri":"P3","acct":"Kyndryl","band":"Gold","ado":"—","raised":"Nagalakshmi N","team":"Ganga Reddy","upd":"Jun 12","status":"In Progress","age":11,"reason":""},
    {"id":"219541","subj":"Trustech - Finops not triggered","pri":"P2","acct":"CoreStack (Internal)","band":"Bronze","ado":"—","raised":"Krishna Kumar VJ","team":"Ganga Reddy","upd":"Jun 12","status":"In Progress","age":6,"reason":""},
    {"id":"216808","subj":"Difference Between CoreStack Recommendation and Azure Calculation","pri":"P3","acct":"Kyndryl","band":"Gold","ado":"132316","raised":"Randhir Kumar","team":"PremKumar B","upd":"Jun 12","status":"In Progress","age":13,"reason":""},
    {"id":"216750","subj":"FW: Production Environment IFoundry5X","pri":"P3","acct":"Tata Communications","band":"Gold","ado":"132646","raised":"Sunilkumar S","team":"PremKumar B","upd":"Jun 12","status":"In Progress","age":13,"reason":""},
    {"id":"220105","subj":"Merged cells in cost recommendation Report","pri":"P3","acct":"Kyndryl","band":"Gold","ado":"133816","raised":"Nagalakshmi N","team":"PremKumar B","upd":"Jun 12","status":"In Progress","age":5,"reason":""},
    {"id":"205888","subj":"Need assistance to update creds for EA account for Trinity College","pri":"P2","acct":"Logicalis","band":"Gold","ado":"132186","raised":"Kamran Wahid","team":"PremKumar B","upd":"Jun 12","status":"In Progress","age":41,"reason":""},
    {"id":"220546","subj":"ODP/Blackstone - Bus Patrol","pri":"P3","acct":"CoreStack (Internal)","band":"Bronze","ado":"—","raised":"Jayven Couch","team":"Nithin Ram","upd":"Jun 12","status":"In Progress","age":4,"reason":""},
    {"id":"220875","subj":"Dashboard slowness","pri":"P2","acct":"Synopsys","band":"Gold","ado":"133875","raised":"Ranjitha Thota","team":"Ganga Reddy","upd":"Jun 14","status":"In Progress","age":3,"reason":""},
    {"id":"221821","subj":"GE Reports not working","pri":"P3","acct":"GE Vernova","band":"Bronze","ado":"133952","raised":"Vijay Kumar P","team":"Avinash Naidu","upd":"Jun 14","status":"In Progress","age":1,"reason":""},
    {"id":"221940","subj":"Re: Corestack Project Addition","pri":"P3","acct":"LTTS","band":"Bronze","ado":"—","raised":"Kaustubh M","team":"Avinash Naidu","upd":"Jun 15","status":"In Progress","age":0,"reason":""},
    {"id":"222003","subj":"SHI Locuz - Need Assistance in Compliance Execution","pri":"P3","acct":"CoreStack (Internal)","band":"Bronze","ado":"—","raised":"Nagalakshmi N","team":"Deepesh H","upd":"Jun 15","status":"In Progress","age":0,"reason":""},
    {"id":"220999","subj":"Request for Investigation – OCI Cost Processing","pri":"P3","acct":"Core42","band":"Bronze","ado":"133976","raised":"Muthu D","team":"PremKumar B","upd":"Jun 15","status":"In Progress","age":3,"reason":""},
    {"id":"222040","subj":"Billing Amount Difference between GCP and Core Stock","pri":"P3","acct":"LTTS","band":"Bronze","ado":"—","raised":"Kaustubh M","team":"Deepesh H","upd":"Jun 15","status":"In Progress","age":0,"reason":""},
    {"id":"206833","subj":"Mar26 usage for Mitsui Chemicals","pri":"P3","acct":"Synoptek","band":"Gold","ado":"129985","raised":"Stacey Zborowski","team":"Nithin Ram","upd":"Jun 15","status":"In Progress","age":39,"reason":""},
    # ON HOLD
    {"id":"209005","subj":"Getting wrong recommended SKU in cost recommendation report","pri":"P2","acct":"Neurealm","band":"Platinum","ado":"130297","raised":"Swapnilyadav Ingale","team":"PremKumar B","upd":"Jun 14","status":"On Hold","age":33,"reason":"The reported invalid recommendation issue has been fixed. Reviewed all the remaining right sizing system recommendations and they appear to be valid. However the customer recently has raised a concern that all the recommendations provided by CS are invalid. So we have given the context and informed Nagalakshmi to reply in this ticket. Hence we are keeping it on hold."},
    {"id":"211893","subj":"Re: Core stock Finops Dashboard cost differ","pri":"P2","acct":"Neurealm","band":"Platinum","ado":"—","raised":"Parthasarathy K","team":"Avinash Naidu","upd":"Jun 11","status":"On Hold","age":25,"reason":"Customer needs to raise the support case with Azure. This is an issue from Azure side."},
    {"id":"211895","subj":"Re: Core stock Finops Dashboard cost differ for GCP","pri":"P3","acct":"Neurealm","band":"Platinum","ado":"131068","raised":"Parthasarathy K","team":"Nithin Ram","upd":"Jun 15","status":"On Hold","age":25,"reason":"Steps provided customer has to implement the changes."},
    {"id":"211954","subj":"AWS Accounts transfer from INH to ISO Tenant","pri":"P3","acct":"Otsuka","band":"Gold","ado":"—","raised":"Rajkumar Uppu","team":"Aadhithya Shanmugapriyan","upd":"Jun 15","status":"On Hold","age":25,"reason":"Awaiting confirmation from Ashok to proceed with backfilling of the cost data for these 3 accounts."},
    {"id":"215451","subj":"Re: Corestack Project Addition","pri":"P3","acct":"LTTS","band":"Bronze","ado":"—","raised":"Kaustubh M","team":"Avinash Naidu","upd":"Jun 14","status":"On Hold","age":17,"reason":"Waiting for the customer to provide the availability so that we can get into a call to discuss this further."},
    {"id":"217785","subj":"RDS Snapshot Not Created on May 31","pri":"P3","acct":"Cloud Kinetics","band":"Silver","ado":"132850","raised":"Service Assurance","team":"Gnanadesigan A","upd":"Jun 11","status":"On Hold","age":10,"reason":"We have stated that we do not have sufficient logs to troubleshoot further and the customer is checking internally."},
    {"id":"217961","subj":"Sonata - CSP accounts not showing up","pri":"P3","acct":"Sonata","band":"Bronze","ado":"132668","raised":"Deovrat Soman","team":"PremKumar B","upd":"Jun 15","status":"On Hold","age":10,"reason":"The initial reported issue has been resolved, however while loading the dashboard we are encountering errors. Engineering team suspects the issue is due to missing currency."},
    {"id":"217989","subj":"ODP - ALiando - CoreTrust - National Tree - Cost Processing","pri":"P3","acct":"CoreStack (Internal)","band":"Bronze","ado":"—","raised":"Anaranya Bagchi","team":"Ganga Reddy","upd":"Jun 12","status":"On Hold","age":10,"reason":"Anaranya has sent a mail to the customer to allow the API permission from CSP Partner."},
    {"id":"217990","subj":"Cloud.corestack.io is slow across all pages","pri":"P3","acct":"CoreStack (Internal)","band":"Bronze","ado":"—","raised":"Satyabrat","team":"Ganga Reddy","upd":"Jun 15","status":"On Hold","age":10,"reason":"We are awaiting response from Pendo team."},
    {"id":"219147","subj":"RE: RE:[CASE] CUR Backfill","pri":"P3","acct":"Sonata","band":"Bronze","ado":"—","raised":"Raghavan P","team":"Nithin Ram","upd":"Jun 13","status":"On Hold","age":7,"reason":"Flow currently being tested and 1 account works as expected. Will proceed with the remaining."},
    {"id":"219360","subj":"cloud.corestack.io","pri":"P3","acct":"CoreStack (Internal)","band":"Bronze","ado":"—","raised":"Satyabrat","team":"Nithin Ram","upd":"Jun 9","status":"On Hold","age":7,"reason":"NA"},
    # AWAITING RESOLUTION CONFIRMATION
    {"id":"214686","subj":"No cost data for TreeRing (AEMCS)","pri":"P3","acct":"CoreStack (Internal)","band":"Bronze","ado":"—","raised":"Jayven Couch","team":"Logesh S","upd":"Jun 14","status":"Awaiting Resolution Confirmation","age":19,"reason":"Waiting for the ticket owner to confirm."},
    {"id":"216586","subj":"Unable to onboard Snowflake in CS4CS","pri":"P3","acct":"CoreStack (Internal)","band":"Bronze","ado":"—","raised":"Anaranya Bagchi","team":"Ganga Reddy","upd":"Jun 12","status":"Awaiting Resolution Confirmation","age":14,"reason":"Waiting for Anaranya's availability for call."},
    {"id":"217606","subj":"Login Issue with CoreStack Tool","pri":"P3","acct":"Otsuka","band":"Gold","ado":"—","raised":"Rajkumar Uppu","team":"Ganga Reddy","upd":"Jun 15","status":"Awaiting Resolution Confirmation","age":11,"reason":"Customer pinged in teams and asked to hold for 1 day."},
    {"id":"219361","subj":"Filtrona Finops Dashboard Unallocated Resource Groups","pri":"P3","acct":"Getronics","band":"Silver","ado":"—","raised":"Shashank Nayakt","team":"Nithin Ram","upd":"Jun 11","status":"Awaiting Resolution Confirmation","age":7,"reason":"NA"},
    {"id":"219377","subj":"ODP Corporation MCA Billing Account cost processing errors","pri":"P3","acct":"CoreStack (Internal)","band":"Bronze","ado":"—","raised":"Jayven Couch","team":"Nithin Ram","upd":"Jun 9","status":"Awaiting Resolution Confirmation","age":7,"reason":"Awaiting credential refresh to validate the cost process."},
    {"id":"219658","subj":"US Prod - Dashboard Not Loading","pri":"P3","acct":"CoreStack (Internal)","band":"Bronze","ado":"—","raised":"Ashok Kumar Elangovan","team":"Nithin Ram","upd":"Jun 10","status":"Awaiting Resolution Confirmation","age":6,"reason":"NA"},
    {"id":"219989","subj":"Re: Corestack","pri":"P3","acct":"Sonata","band":"Bronze","ado":"—","raised":"Deovrat Soman","team":"PremKumar B","upd":"Jun 14","status":"Awaiting Resolution Confirmation","age":5,"reason":"Unable to reproduce the issue, informed the same to the customer and we are awaiting their response."},
    {"id":"220165","subj":"Trustedtech - HMH - cost for April","pri":"P3","acct":"CoreStack (Internal)","band":"Bronze","ado":"—","raised":"Krishna Kumar VJ","team":"Nithin Ram","upd":"Jun 10","status":"Awaiting Resolution Confirmation","age":5,"reason":"N/A"},
    {"id":"220416","subj":"Sonata - Tata Tele CSP processing issue","pri":"P2","acct":"Sonata","band":"Bronze","ado":"133815","raised":"Deovrat Soman","team":"PremKumar B","upd":"Jun 15","status":"Awaiting Resolution Confirmation","age":4,"reason":"Cost has been processed and we are now awaiting customer's confirmation."},
    {"id":"220456","subj":"US SaaS - Kyndryl Lifelabs - 2 subscriptions are not available","pri":"P3","acct":"Kyndryl","band":"Gold","ado":"—","raised":"Nagalakshmi N","team":"Avinash Naidu","upd":"Jun 15","status":"Awaiting Resolution Confirmation","age":4,"reason":"NA"},
    {"id":"221054","subj":"US SaaS - Kyndryl - Default dashboards not visible","pri":"P3","acct":"Kyndryl","band":"Gold","ado":"—","raised":"Nagalakshmi N","team":"Nithin Ram","upd":"Jun 12","status":"Awaiting Resolution Confirmation","age":3,"reason":"NA"},
    # OPEN
    {"id":"219258","subj":"Firing: High Priority MSProd App Server Memory Utilisation above 90%","pri":"Normal","acct":"CoreStack (Internal)","band":"Bronze","ado":"—","raised":"Notify SRE Ops","team":"—","upd":"Jun 8","status":"Open","age":7,"reason":""},
    {"id":"219898","subj":"Deployment Status Confirmation Required","pri":"P3","acct":"Cloud Kinetics","band":"Silver","ado":"—","raised":"Service Assurance","team":"PremKumar B","upd":"Jun 10","status":"Open","age":5,"reason":""},
    {"id":"220491","subj":"FW: Resources Cost - Beside Tagged and untagged - LifeLabs","pri":"P3","acct":"Kyndryl","band":"Gold","ado":"133656","raised":"Nagalakshmi N","team":"Avinash Naidu","upd":"Jun 12","status":"Open","age":4,"reason":""},
]

# ── HELPERS ─────────────────────────────────────────────────────────────────

def age_bucket(age):
    if age >= 30: return "30d+"
    if age >= 15: return "15-30d"
    if age >= 8:  return "8-14d"
    return "0-7d"

def age_label(age):
    return f"{age}d"

def l1l2(ado):
    return "L2" if ado and ado != "—" else "L1"

# ── HTML GENERATION ─────────────────────────────────────────────────────────

def pri_badge(p):
    if p == "P2":
        return f'<span style="background:#FFF7ED;color:#C2410C;border-radius:4px;padding:2px 7px;font-size:11px;font-weight:600;">{p}</span>'
    elif p == "P3":
        return f'<span style="background:#F1F5F9;color:#475569;border-radius:4px;padding:2px 7px;font-size:11px;font-weight:600;">{p}</span>'
    else:
        return f'<span style="background:#F1F5F9;color:#475569;border-radius:4px;padding:2px 7px;font-size:11px;font-weight:600;">{p}</span>'

def status_badge(s):
    styles = {
        "In Progress": ("F0FDF4","15803D"),
        "On Hold": ("FFFBEB","B45309"),
        "Awaiting Resolution Confirmation": ("EDE9FE","6D28D9"),
        "Open": ("EFF6FF","1D4ED8"),
    }
    bg, fg = styles.get(s, ("F1F5F9","475569"))
    label = "ARC" if s == "Awaiting Resolution Confirmation" else s
    return f'<span style="background:#{bg};color:#{fg};border-radius:4px;padding:2px 7px;font-size:11px;font-weight:600;white-space:nowrap;">{label}</span>'

def age_badge(age):
    b = age_bucket(age)
    colors = {
        "30d+":  ("FEE2E2","B91C1C"),
        "15-30d":("FEF3C7","B45309"),
        "8-14d": ("FEFCE8","713F12"),
        "0-7d":  ("F0FDF4","14532D"),
    }
    bg, fg = colors[b]
    return f'<span style="background:#{bg};color:#{fg};border-radius:4px;padding:2px 7px;font-size:11px;font-weight:600;">{age}d</span>'

def team_badge(ado):
    if ado and ado != "—":
        return '<span style="background:#EDE9FE;color:#6D28D9;border-radius:4px;padding:2px 7px;font-size:11px;font-weight:700;">L2</span>'
    return '<span style="background:#F1F5F9;color:#64748B;border-radius:4px;padding:2px 7px;font-size:11px;font-weight:600;">L1</span>'

def ado_badge(ado):
    if ado and ado != "—":
        return f'<span style="background:#F5F3FF;color:#6D28D9;border-radius:4px;padding:2px 7px;font-size:11px;font-weight:600;">{ado}</span>'
    return ''

def row_bg(t):
    s, a = t["status"], t["age"]
    if s == "On Hold" and a >= 30: return "#FFF0F0"
    if s == "On Hold" and a >= 8:  return "#FFFBF0"
    if s == "In Progress" and 8 <= a <= 14: return "#FEFFF0"
    return "#FFFFFF"

def band_badge(band):
    styles = {
        "Platinum": 'background:#1E293B;color:#CBD5E1;border:1px solid #334155;',
        "Gold":     'background:#FFFBEB;color:#92400E;border:1px solid #FDE68A;',
        "Silver":   'background:#F1F5F9;color:#475569;border:1px solid #CBD5E1;',
        "Bronze":   'background:#FFF7ED;color:#9A3412;border:1px solid #FED7AA;',
    }
    st = styles.get(band, '')
    return f'<span style="{st}border-radius:4px;padding:2px 8px;font-size:10px;font-weight:700;">{band}</span>'

def status_pill(label, color):
    colors = {
        "green":  ("F0FDF4","15803D"),
        "amber":  ("FFFBEB","B45309"),
        "purple": ("EDE9FE","6D28D9"),
        "blue":   ("EFF6FF","1D4ED8"),
        "red":    ("FEE2E2","B91C1C"),
    }
    bg, fg = colors.get(color, ("F1F5F9","475569"))
    return f'<span style="background:#{bg};color:#{fg};border-radius:10px;padding:2px 10px;font-size:11px;font-weight:600;margin-right:4px;">{label}</span>'

def account_section(acct, band, tickets_in):
    # count pills
    from collections import Counter
    sc = Counter()
    for t in tickets_in:
        sc[t["status"]] += 1
    pills_html = ""
    if sc.get("In Progress"):
        pills_html += status_pill(f'{sc["In Progress"]} In Progress', "green")
    if sc.get("On Hold"):
        pills_html += status_pill(f'{sc["On Hold"]} On Hold', "amber")
    if sc.get("Awaiting Resolution Confirmation"):
        pills_html += status_pill(f'{sc["Awaiting Resolution Confirmation"]} Awaiting Resolution Confirmation', "purple")
    if sc.get("Open"):
        pills_html += status_pill(f'{sc["Open"]} Open', "blue")

    rows_html = ""
    for t in tickets_in:
        bg = row_bg(t)
        rows_html += f"""
        <tr style="background:{bg};border-bottom:1px solid #F1F5F9;">
          <td style="padding:8px 10px;"><a href="#" style="color:#2563EB;font-weight:700;text-decoration:none;">#{t['id']}</a></td>
          <td style="padding:8px 10px;font-size:12px;">{t['subj']}</td>
          <td style="padding:8px 10px;">{pri_badge(t['pri'])}</td>
          <td style="padding:8px 10px;">{status_badge(t['status'])}</td>
          <td style="padding:8px 10px;">{age_badge(t['age'])}</td>
          <td style="padding:8px 10px;">{team_badge(t['ado'])}</td>
          <td style="padding:8px 10px;">{ado_badge(t['ado'])}</td>
          <td style="padding:8px 10px;font-size:12px;">{t['raised']}</td>
          <td style="padding:8px 10px;font-size:12px;">{t['upd']}</td>
        </tr>"""

    return f"""
    <tr><td height="18"></td></tr>
    <tr><td>
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-radius:10px;overflow:hidden;border:1px solid #E4E8EF;">
        <tr style="background:#F8FAFC;">
          <td style="padding:12px 16px;">
            <span style="font-size:14px;font-weight:700;color:#0F172A;">{acct}</span>
            &nbsp;&nbsp;{band_badge(band)}&nbsp;&nbsp;
            {pills_html}
          </td>
        </tr>
        <tr><td>
          <table width="100%" cellpadding="0" cellspacing="0" border="0">
            <tr style="background:#F8FAFC;">
              <td style="width:7%;padding:7px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;">Ticket</td>
              <td style="width:22%;padding:7px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;">Subject</td>
              <td style="width:8%;padding:7px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;">Priority</td>
              <td style="width:13%;padding:7px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;">Status</td>
              <td style="width:5%;padding:7px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;">Age</td>
              <td style="width:5%;padding:7px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;">Team</td>
              <td style="width:6%;padding:7px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;">ADO #</td>
              <td style="width:13%;padding:7px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;">Raised By</td>
              <td style="width:13%;padding:7px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;">Last Update</td>
            </tr>
            {rows_html}
          </table>
        </td></tr>
      </table>
    </td></tr>"""

# Account sections order
ACCOUNT_ORDER = [
    ("Neurealm",           "Platinum"),
    ("Kyndryl",            "Gold"),
    ("Logicalis",          "Gold"),
    ("Otsuka",             "Gold"),
    ("Synopsys",           "Gold"),
    ("Synoptek",           "Gold"),
    ("Tata Communications","Gold"),
    ("Cloud Kinetics",     "Silver"),
    ("Getronics",          "Silver"),
]

SHOW_BANDS = {"Platinum","Gold","Silver"}

def build_html():
    # counts
    new_c   = 3
    ip_c    = 15
    oh_c    = 11
    arc_c   = 11
    l2_c    = 14

    aging_30  = 3
    aging_15  = 6
    aging_8   = 5
    aging_0   = 26

    plat_c   = 3
    gold_c   = 13
    silver_c = 3

    acct_sections = ""
    for acct, band in ACCOUNT_ORDER:
        tickets_in = [t for t in TICKETS if t["acct"] == acct and t["band"] == band]
        acct_sections += account_section(acct, band, tickets_in)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><title>Daily Incident Report &middot; June 15, 2026</title></head>
<body style="margin:0;padding:0;background:#F4F6F9;font-family:Arial,sans-serif;font-size:13px;color:#1A2035;">
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#F4F6F9;"><tr><td align="center" style="padding:20px 16px 40px;">
<table width="980" cellpadding="0" cellspacing="0" border="0" style="max-width:980px;width:100%;">

<!-- HEADER -->
<tr><td style="background:#fff;border:1px solid #E4E8EF;border-radius:10px;padding:18px 28px 16px;">
  <div style="font-size:17px;font-weight:700;color:#0F172A;">Daily Incident Report</div>
  <div style="font-size:11px;color:#64748B;margin-top:3px;"><b style="color:#334155;">Period:</b>&nbsp;June 15, 2026 &middot; 19:30 IST&nbsp;&nbsp;<b style="color:#334155;">Dept:</b>&nbsp;CoreStack Support</div>
</td></tr>
<tr><td height="14"></td></tr>

<!-- SUMMARY -->
<tr><td style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:#94A3B8;padding-bottom:10px;">Summary</td></tr>
<tr><td>
  <table width="100%" cellpadding="0" cellspacing="0" border="0"><tr>
    <td width="20%" style="padding-right:10px;">
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#fff;border:1px solid #E4E8EF;border-radius:10px;border-top:3px solid #3B82F6;">
        <tr><td style="padding:12px 14px;">
          <div style="font-size:11px;color:#64748B;font-weight:600;">New</div>
          <div style="font-size:28px;font-weight:700;color:#1A2035;">{new_c}</div>
          <div style="font-size:10px;color:#94A3B8;">opened</div>
        </td></tr>
      </table>
    </td>
    <td width="20%" style="padding-right:10px;">
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#fff;border:1px solid #E4E8EF;border-radius:10px;border-top:3px solid #10B981;">
        <tr><td style="padding:12px 14px;">
          <div style="font-size:11px;color:#64748B;font-weight:600;">In Progress</div>
          <div style="font-size:28px;font-weight:700;color:#1A2035;">{ip_c}</div>
          <div style="font-size:10px;color:#94A3B8;">being worked</div>
        </td></tr>
      </table>
    </td>
    <td width="20%" style="padding-right:10px;">
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#fff;border:1px solid #E4E8EF;border-radius:10px;border-top:3px solid #F59E0B;">
        <tr><td style="padding:12px 14px;">
          <div style="font-size:11px;color:#64748B;font-weight:600;">On Hold</div>
          <div style="font-size:28px;font-weight:700;color:#1A2035;">{oh_c}</div>
          <div style="font-size:10px;color:#94A3B8;">pending / monitoring</div>
        </td></tr>
      </table>
    </td>
    <td width="20%" style="padding-right:10px;">
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#fff;border:1px solid #E4E8EF;border-radius:10px;border-top:3px solid #8B5CF6;">
        <tr><td style="padding:12px 14px;">
          <div style="font-size:11px;color:#64748B;font-weight:600;">Awaiting Confirmation</div>
          <div style="font-size:28px;font-weight:700;color:#1A2035;">{arc_c}</div>
          <div style="font-size:10px;color:#94A3B8;">awaiting customer</div>
        </td></tr>
      </table>
    </td>
    <td width="20%">
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#fff;border:1px solid #E4E8EF;border-radius:10px;border-top:3px solid #EF4444;">
        <tr><td style="padding:12px 14px;">
          <div style="font-size:11px;color:#64748B;font-weight:600;">With L2 (PGS)</div>
          <div style="font-size:28px;font-weight:700;color:#1A2035;">{l2_c}</div>
          <div style="font-size:10px;color:#94A3B8;">ADO linked</div>
        </td></tr>
      </table>
    </td>
  </tr></table>
</td></tr>
<tr><td height="18"></td></tr>

<!-- AGING -->
<tr><td style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:#94A3B8;padding-bottom:10px;">Open Ticket Aging</td></tr>
<tr><td>
  <table width="100%" cellpadding="0" cellspacing="0" border="0"><tr>
    <td width="25%" style="padding-right:10px;">
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#FEF2F2;border:1px solid #FECACA;border-radius:8px;">
        <tr><td style="padding:10px 14px;">
          <div style="font-size:11px;font-weight:600;color:#B91C1C;">30d+</div>
          <div style="font-size:26px;font-weight:700;color:#DC2626;text-align:right;">{aging_30}</div>
        </td></tr>
      </table>
    </td>
    <td width="25%" style="padding-right:10px;">
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#FFFBEB;border:1px solid #FDE68A;border-radius:8px;">
        <tr><td style="padding:10px 14px;">
          <div style="font-size:11px;font-weight:600;color:#B45309;">15&ndash;30d</div>
          <div style="font-size:26px;font-weight:700;color:#D97706;text-align:right;">{aging_15}</div>
        </td></tr>
      </table>
    </td>
    <td width="25%" style="padding-right:10px;">
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#FEFCE8;border:1px solid #FEF08A;border-radius:8px;">
        <tr><td style="padding:10px 14px;">
          <div style="font-size:11px;font-weight:600;color:#CA8A04;">8&ndash;14d</div>
          <div style="font-size:26px;font-weight:700;color:#CA8A04;text-align:right;">{aging_8}</div>
        </td></tr>
      </table>
    </td>
    <td width="25%">
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#F0FDF4;border:1px solid #BBF7D0;border-radius:8px;">
        <tr><td style="padding:10px 14px;">
          <div style="font-size:11px;font-weight:600;color:#16A34A;">0&ndash;7d</div>
          <div style="font-size:26px;font-weight:700;color:#16A34A;text-align:right;">{aging_0}</div>
        </td></tr>
      </table>
    </td>
  </tr></table>
</td></tr>
<tr><td height="18"></td></tr>

<!-- BAND COUNTS -->
<tr><td style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:#94A3B8;padding-bottom:10px;">Tickets by Account Band</td></tr>
<tr><td>
  <table width="100%" cellpadding="0" cellspacing="0" border="0"><tr>
    <td width="33%" style="padding-right:10px;">
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#1E293B;border:1px solid #334155;border-radius:10px;">
        <tr><td style="padding:14px 16px;">
          <div style="font-size:12px;font-weight:600;color:#CBD5E1;">Platinum</div>
          <div style="font-size:28px;font-weight:700;color:#CBD5E1;text-align:right;">{plat_c}</div>
        </td></tr>
      </table>
    </td>
    <td width="33%" style="padding-right:10px;">
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#451A03;border:1px solid #78350F;border-radius:10px;">
        <tr><td style="padding:14px 16px;">
          <div style="font-size:12px;font-weight:600;color:#FBBF24;">Gold</div>
          <div style="font-size:28px;font-weight:700;color:#FBBF24;text-align:right;">{gold_c}</div>
        </td></tr>
      </table>
    </td>
    <td width="33%">
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#1E3A5F;border:1px solid #1E4976;border-radius:10px;">
        <tr><td style="padding:14px 16px;">
          <div style="font-size:12px;font-weight:600;color:#93C5FD;">Silver</div>
          <div style="font-size:28px;font-weight:700;color:#93C5FD;text-align:right;">{silver_c}</div>
        </td></tr>
      </table>
    </td>
  </tr></table>
</td></tr>
<tr><td height="18"></td></tr>

<!-- LEGEND -->
<tr><td style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:#94A3B8;padding-bottom:10px;">Ticket Detail by Account (Platinum &middot; Gold &middot; Silver)</td></tr>
<tr><td>
  <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#F8FAFC;border:1px solid #E4E8EF;border-radius:10px;">
    <tr>
      <td width="50%" style="padding:14px 20px;border-right:1px solid #E4E8EF;">
        <div style="font-size:12px;font-weight:700;color:#B45309;">&#x1F7E1; On Hold</div>
        <div style="font-size:11px;color:#64748B;margin-top:4px;">Ticket is blocked pending customer action, third-party response, or an internal dependency. No further progress can be made by the support team until the blocker is resolved.</div>
      </td>
      <td width="50%" style="padding:14px 20px 14px 24px;">
        <div style="font-size:12px;font-weight:700;color:#6D28D9;">&#x1F7E3; Awaiting Resolution Confirmation</div>
        <div style="font-size:11px;color:#64748B;margin-top:4px;">A fix or workaround has been applied and communicated to the customer. The ticket remains open until the customer confirms the issue is resolved and the case can be formally closed.</div>
      </td>
    </tr>
  </table>
</td></tr>

<!-- ACCOUNT SECTIONS -->
{acct_sections}

<tr><td height="20"></td></tr>

<!-- FOOTER -->
<tr><td>
  <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#F0F9FF;border:1px solid #BAE6FD;border-radius:10px;">
    <tr><td style="padding:14px 20px;">
      <span style="font-size:11px;color:#64748B;">&#x1F4CA; Full ticket dump (Excel): </span>
      <a href="CS_Daily_Incident_Report_20260615.xlsx" style="font-size:11px;font-weight:600;color:#2563EB;text-decoration:none;">Download Daily Incident Report Excel &rarr;</a>
    </td></tr>
  </table>
</td></tr>

</table>
</td></tr></table>
</body>
</html>"""
    return html

html_path = "/home/user/GD/CS_Daily_Incident_Report_20260615.html"
with open(html_path, "w", encoding="utf-8") as f:
    f.write(build_html())
print(f"HTML written: {html_path}")

# ── EXCEL GENERATION ─────────────────────────────────────────────────────────

try:
    import openpyxl
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── style helpers ──

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))

def font(color="#000000", bold=False, size=11):
    return Font(color=color.lstrip("#"), bold=bold, size=size)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

BAND_FILL = {
    "Platinum": ("CBD5E1","1E293B"),
    "Gold":     ("FEF9C3","92400E"),
    "Silver":   ("DBEAFE","1E4976"),
    "Bronze":   ("FFF7ED","9A3412"),
}
STATUS_FILL = {
    "Open":                           ("DBEAFE","1D4ED8"),
    "In Progress":                    ("DCFCE7","15803D"),
    "On Hold":                        ("FEF9C3","B45309"),
    "Awaiting Resolution Confirmation":("EDE9FE","6D28D9"),
}

def write_header(ws, headers, hdr_fill, hdr_font_color="FFFFFF"):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = fill(hdr_fill)
        cell.font = font(f"#{hdr_font_color}", bold=True, size=10)
        cell.alignment = center()
        cell.border = border
    ws.row_dimensions[1].height = 20

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def style_data_row(ws, row_num, t, has_reason=True):
    row = ws[row_num]
    bf, ff = BAND_FILL.get(t["band"], ("FFFFFF","000000"))
    sf, sff = STATUS_FILL.get(t["status"], ("FFFFFF","000000"))
    for cell in row:
        cell.border = border
        cell.alignment = left()
        cell.font = font("#1A2035", size=10)
    ws.row_dimensions[row_num].height = 30
    # Band col (index 5 for 14-col = E, index 5 for 13-col = E)
    # We'll color the status cell specially
    # Col order: Ticket# | Subject | Priority | Account | Band | Assignee | Status | Age | AgeBucket | L1L2 | ADO | RaisedBy | LastUpdate [| Reason]
    # indices:     1         2         3          4         5      6          7        8     9           10     11    12          13          14
    band_col = 5
    status_col = 7
    row[band_col-1].fill = fill(f"#{bf}")
    row[band_col-1].font = font(f"#{ff}", bold=True, size=10)
    row[status_col-1].fill = fill(f"#{sf}")
    row[status_col-1].font = font(f"#{sff}", bold=True, size=10)

def build_ticket_row(t, has_reason=True):
    row = [
        t["id"],
        t["subj"],
        t["pri"],
        t["acct"],
        t["band"],
        t["team"],
        t["status"],
        t["age"],
        age_bucket(t["age"]),
        l1l2(t["ado"]),
        t["ado"] if t["ado"] != "—" else "",
        t["raised"],
        t["upd"],
    ]
    if has_reason:
        row.append(t.get("reason",""))
    return row

HDR14 = ["Ticket #","Subject","Priority","Account","Band","Assignee","Status","Age (days)","Age Bucket","L1/L2","ADO Link","Raised By","Last Update","Hold/ARC Reason"]
HDR13 = ["Ticket #","Subject","Priority","Account","Band","Assignee","Status","Age (days)","Age Bucket","L1/L2","ADO Link","Raised By","Last Update"]

WIDTHS14 = [10,40,10,26,8,18,18,10,10,8,12,18,14,50]
WIDTHS13 = [10,40,10,26,8,18,18,10,10,8,12,18,14]

# ── 1. Summary ──
ws1 = wb.active
ws1.title = "Summary"
ws1.append(["Status","Count","%"])
for cell in ws1[1]:
    cell.fill = fill("1E293B"); cell.font = font("#FFFFFF", bold=True); cell.alignment = center(); cell.border = border
ws1.row_dimensions[1].height = 20

summary_data = [
    ("Open",3,"7.5%"),
    ("In Progress",15,"37.5%"),
    ("On Hold",11,"27.5%"),
    ("Awaiting Resolution Confirmation",11,"27.5%"),
    ("Total",40,"100%"),
]
for i,(s,c,p) in enumerate(summary_data, 2):
    ws1.append([s,c,p])
    row = ws1[i]
    for cell in row:
        cell.border = border; cell.alignment = left(); cell.font = font("#1A2035",size=10)
    ws1.row_dimensions[i].height = 30
    if s in STATUS_FILL:
        sf,sff = STATUS_FILL[s]
        row[0].fill = fill(f"#{sf}"); row[0].font = font(f"#{sff}",bold=True,size=10)
    elif s == "Total":
        for cell in row:
            cell.font = font("#1A2035",bold=True,size=10)
ws1.column_dimensions["A"].width = 32
ws1.column_dimensions["B"].width = 10
ws1.column_dimensions["C"].width = 10

# ── 2. All Tickets ──
ws2 = wb.create_sheet("All Tickets")
write_header(ws2, HDR14, "1E293B")
set_col_widths(ws2, WIDTHS14)
for i,t in enumerate(TICKETS, 2):
    ws2.append(build_ticket_row(t, True))
    style_data_row(ws2, i, t, True)

# ── 3. Platinum Gold Silver ──
ws3 = wb.create_sheet("Platinum Gold Silver")
write_header(ws3, HDR14, "334155")
set_col_widths(ws3, WIDTHS14)
pgs = [t for t in TICKETS if t["band"] in SHOW_BANDS]
for i,t in enumerate(pgs, 2):
    ws3.append(build_ticket_row(t, True))
    style_data_row(ws3, i, t, True)

# ── 4. New Open ──
ws4 = wb.create_sheet("New Open")
write_header(ws4, HDR13, "1D4ED8")
set_col_widths(ws4, WIDTHS13)
opens = [t for t in TICKETS if t["status"] == "Open"]
for i,t in enumerate(opens, 2):
    ws4.append(build_ticket_row(t, False))
    style_data_row(ws4, i, t, False)

# ── 5. In Progress ──
ws5 = wb.create_sheet("In Progress")
write_header(ws5, HDR13, "15803D")
set_col_widths(ws5, WIDTHS13)
ip = [t for t in TICKETS if t["status"] == "In Progress"]
for i,t in enumerate(ip, 2):
    ws5.append(build_ticket_row(t, False))
    style_data_row(ws5, i, t, False)

# ── 6. On Hold ──
ws6 = wb.create_sheet("On Hold")
write_header(ws6, HDR14, "B45309")
set_col_widths(ws6, WIDTHS14)
oh = [t for t in TICKETS if t["status"] == "On Hold"]
for i,t in enumerate(oh, 2):
    ws6.append(build_ticket_row(t, True))
    style_data_row(ws6, i, t, True)

# ── 7. Awaiting Resolution ──
ws7 = wb.create_sheet("Awaiting Resolution")
write_header(ws7, HDR14, "6D28D9")
set_col_widths(ws7, WIDTHS14)
arc = [t for t in TICKETS if t["status"] == "Awaiting Resolution Confirmation"]
for i,t in enumerate(arc, 2):
    ws7.append(build_ticket_row(t, True))
    style_data_row(ws7, i, t, True)

# ── 8. L2 Tickets ──
ws8 = wb.create_sheet("L2 Tickets")
write_header(ws8, HDR14, "6D28D9")
set_col_widths(ws8, WIDTHS14)
l2 = [t for t in TICKETS if t["ado"] and t["ado"] != "—"]
for i,t in enumerate(l2, 2):
    ws8.append(build_ticket_row(t, True))
    style_data_row(ws8, i, t, True)

# ── 9. Aging View ──
ws9 = wb.create_sheet("Aging View")
AGE_HDR = ["Age Bucket","Ticket #","Subject","Priority","Account","Band","Assignee","Status"]
ws9.append(AGE_HDR)
for cell in ws9[1]:
    cell.fill = fill("334155"); cell.font = font("#FFFFFF",bold=True); cell.alignment = center(); cell.border = border
ws9.row_dimensions[1].height = 20
ws9.column_dimensions["A"].width = 16
ws9.column_dimensions["B"].width = 10
ws9.column_dimensions["C"].width = 44
ws9.column_dimensions["D"].width = 10
ws9.column_dimensions["E"].width = 26
ws9.column_dimensions["F"].width = 10
ws9.column_dimensions["G"].width = 20
ws9.column_dimensions["H"].width = 20

BUCKET_LABELS = [
    ("30d+",  [t for t in TICKETS if age_bucket(t["age"])=="30d+"]),
    ("15-30d",[t for t in TICKETS if age_bucket(t["age"])=="15-30d"]),
    ("8-14d", [t for t in TICKETS if age_bucket(t["age"])=="8-14d"]),
    ("0-7d",  [t for t in TICKETS if age_bucket(t["age"])=="0-7d"]),
]
BUCKET_FILL = {"30d+":"FEE2E2","15-30d":"FEF3C7","8-14d":"FEFCE8","0-7d":"F0FDF4"}
BUCKET_FONT = {"30d+":"B91C1C","15-30d":"B45309","8-14d":"713F12","0-7d":"14532D"}

row_idx = 2
for bname, btickets in BUCKET_LABELS:
    # bucket header row
    ws9.append([f"{bname} ({len(btickets)} tickets)","","","","","","",""])
    hrow = ws9[row_idx]
    for cell in hrow:
        cell.fill = fill(BUCKET_FILL[bname])
        cell.font = font(f"#{BUCKET_FONT[bname]}", bold=True, size=11)
        cell.border = border
        cell.alignment = left()
    ws9.row_dimensions[row_idx].height = 22
    row_idx += 1
    for t in btickets:
        ws9.append([bname, t["id"], t["subj"], t["pri"], t["acct"], t["band"], t["team"], t["status"]])
        drow = ws9[row_idx]
        for cell in drow:
            cell.border = border; cell.alignment = left(); cell.font = font("#1A2035",size=10)
        ws9.row_dimensions[row_idx].height = 30
        bf,ff = BAND_FILL.get(t["band"],("FFFFFF","000000"))
        drow[5].fill = fill(f"#{bf}"); drow[5].font = font(f"#{ff}",bold=True,size=10)
        sf,sff = STATUS_FILL.get(t["status"],("FFFFFF","000000"))
        drow[7].fill = fill(f"#{sf}"); drow[7].font = font(f"#{sff}",bold=True,size=10)
        row_idx += 1

xlsx_path = "/home/user/GD/CS_Daily_Incident_Report_20260615.xlsx"
wb.save(xlsx_path)
print(f"XLSX written: {xlsx_path}")
