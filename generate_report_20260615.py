#!/usr/bin/env python3
"""Generate CS Daily Incident Report for June 15, 2026 - HTML and Excel."""

import subprocess
import sys

# Ensure openpyxl is available
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# DATA
# ──────────────────────────────────────────────

# Fields: ticket, priority, date, age_days, account, band, ado, subject, raised_by, assignee, last_update, status, reason
IN_PROGRESS = [
    (211192,"P3","2026-05-19",27,"Tata Communications","Gold","125029","FinOps Email Notification Issue (Clone)","Sunilkumar S","PremKumar B","Jun 9","In Progress",""),
    (217515,"P3","2026-06-04",11,"Kyndryl","Gold","—","FW: Cost Mismatch in recommendation","Nagalakshmi N","Ganga Reddy","Jun 12","In Progress",""),
    (219541,"P2","2026-06-09",6,"CoreStack (Internal)","Bronze","—","Trustech - Finops not triggered","Krishna Kumar VJ","Ganga Reddy","Jun 12","In Progress",""),
    (216808,"P3","2026-06-02",13,"Kyndryl","Gold","132316","Difference Between CoreStack Recommendation and Azure Calculation","Randhir Kumar","PremKumar B","Jun 12","In Progress",""),
    (216750,"P3","2026-06-02",13,"Tata Communications","Gold","132646","FW: Production Environment IFoundry5X","Sunilkumar S","PremKumar B","Jun 12","In Progress",""),
    (220105,"P3","2026-06-10",5,"Kyndryl","Gold","133816","Merged cells in cost recommendation Report","Nagalakshmi N","PremKumar B","Jun 12","In Progress",""),
    (205888,"P2","2026-05-05",41,"Logicalis","Gold","132186","Need assistance to update creds for EA account for Trinity College","Kamran Wahid","PremKumar B","Jun 12","In Progress",""),
    (220546,"P3","2026-06-11",4,"CoreStack (Internal)","Bronze","—","ODP/Blackstone - Bus Patrol","Jayven Couch","Nithin Ram","Jun 12","In Progress",""),
    (220875,"P2","2026-06-12",3,"Synopsys","Gold","133875","Dashboard slowness","Ranjitha Thota","Ganga Reddy","Jun 14","In Progress",""),
    (221821,"P3","2026-06-14",1,"GE Vernova","Bronze","133952","GE Reports not working","Vijay Kumar P","Avinash Naidu","Jun 14","In Progress",""),
    (221940,"P3","2026-06-15",0,"LTTS","Bronze","—","Re: Corestack Project Addition","Kaustubh M","Avinash Naidu","Jun 15","In Progress",""),
    (222003,"P3","2026-06-15",0,"CoreStack (Internal)","Bronze","—","SHI Locuz - Need Assistance in Compliance Execution","Nagalakshmi N","Deepesh H","Jun 15","In Progress",""),
    (220999,"P3","2026-06-12",3,"Core42","Bronze","133976","Request for Investigation – OCI Cost Processing","Muthu D","PremKumar B","Jun 15","In Progress",""),
    (222040,"P3","2026-06-15",0,"LTTS","Bronze","—","Billing Amount Difference between GCP and Core Stock","Kaustubh M","Deepesh H","Jun 15","In Progress",""),
    (206833,"P3","2026-05-07",39,"Synoptek","Gold","129985","Mar26 usage for Mitsui Chemicals","Stacey Zborowski","Nithin Ram","Jun 15","In Progress",""),
]

ON_HOLD = [
    (209005,"P2","2026-05-13",33,"Neurealm","Platinum","130297","Getting wrong recommended SKU in cost recommendation report","Swapnilyadav Ingale","PremKumar B","Jun 14","On Hold","The reported invalid recommendation issue has been fixed. Reviewed all the remaining right sizing system recommendations and they appear to be valid. However the customer recently has raised a concern that all the recommendations provided by CS are invalid. So we have given the context and informed Nagalakshmi to reply in this ticket. Hence we are keeping it on hold."),
    (211893,"P2","2026-05-21",25,"Neurealm","Platinum","—","Re: Core stock Finops Dashboard cost differ","Parthasarathy K","Avinash Naidu","Jun 11","On Hold","Customer needs to raise the support case with Azure. This is an issue from Azure side."),
    (211895,"P3","2026-05-21",25,"Neurealm","Platinum","131068","Re: Core stock Finops Dashboard cost differ for GCP","Parthasarathy K","Nithin Ram","Jun 15","On Hold","Steps provided customer has to implement the changes."),
    (211954,"P3","2026-05-21",25,"Otsuka","Gold","—","AWS Accounts transfer from INH to ISO Tenant","Rajkumar Uppu","Aadhithya Shanmugapriyan","Jun 15","On Hold","Awaiting confirmation from Ashok to proceed with backfilling of the cost data for these 3 accounts."),
    (215451,"P3","2026-05-29",17,"LTTS","Bronze","—","Re: Corestack Project Addition","Kaustubh M","Avinash Naidu","Jun 14","On Hold","Waiting for the customer to provide the availability so that we can get into a call to discuss this further."),
    (217785,"P3","2026-06-05",10,"Cloud Kinetics","Silver","132850","RDS Snapshot Not Created on May 31","Service Assurance","Gnanadesigan A","Jun 11","On Hold","We have stated that we do not have sufficient logs to troubleshoot further and the customer is checking internally."),
    (217961,"P3","2026-06-05",10,"Sonata","Bronze","132668","Sonata - CSP accounts not showing up","Deovrat Soman","PremKumar B","Jun 15","On Hold","The initial reported issue has been resolved, however while loading the dashboard we are encountering errors. Engineering team suspects the issue is due to missing currency."),
    (217989,"P3","2026-06-05",10,"CoreStack (Internal)","Bronze","—","ODP - ALiando - CoreTrust - National Tree - Cost Processing","Anaranya Bagchi","Ganga Reddy","Jun 12","On Hold","Anaranya has sent a mail to the customer to allow the API permission from CSP Partner."),
    (217990,"P3","2026-06-05",10,"CoreStack (Internal)","Bronze","—","Cloud.corestack.io is slow across all pages","Satyabrat","Ganga Reddy","Jun 15","On Hold","We are awaiting response from Pendo team."),
    (219147,"P3","2026-06-08",7,"Sonata","Bronze","—","RE: RE:[CASE] CUR Backfill","Raghavan P","Nithin Ram","Jun 13","On Hold","Flow currently being tested and 1 account works as expected. Will proceed with the remaining."),
    (219360,"P3","2026-06-08",7,"CoreStack (Internal)","Bronze","—","cloud.corestack.io","Satyabrat","Nithin Ram","Jun 9","On Hold","NA"),
]

ARC = [
    (214686,"P3","2026-05-27",19,"CoreStack (Internal)","Bronze","—","No cost data for TreeRing (AEMCS)","Jayven Couch","Logesh S","Jun 14","Awaiting Resolution Confirmation","Waiting for the ticket owner to confirm."),
    (216586,"P3","2026-06-01",14,"CoreStack (Internal)","Bronze","—","Unable to onboard Snowflake in CS4CS","Anaranya Bagchi","Ganga Reddy","Jun 12","Awaiting Resolution Confirmation","Waiting for Anaranya's availability for call."),
    (217606,"P3","2026-06-04",11,"Otsuka","Gold","—","Login Issue with CoreStack Tool","Rajkumar Uppu","Ganga Reddy","Jun 15","Awaiting Resolution Confirmation","Customer pinged in teams and asked to hold for 1 day."),
    (219361,"P3","2026-06-08",7,"Getronics","Silver","—","Filtrona Finops Dashboard Unallocated Resource Groups","Shashank Nayakt","Nithin Ram","Jun 11","Awaiting Resolution Confirmation","NA"),
    (219377,"P3","2026-06-08",7,"CoreStack (Internal)","Bronze","—","ODP Corporation MCA Billing Account cost processing errors","Jayven Couch","Nithin Ram","Jun 9","Awaiting Resolution Confirmation","Awaiting credential refresh to validate the cost process."),
    (219658,"P3","2026-06-09",6,"CoreStack (Internal)","Bronze","—","US Prod - Dashboard Not Loading","Ashok Kumar Elangovan","Nithin Ram","Jun 10","Awaiting Resolution Confirmation","NA"),
    (219989,"P3","2026-06-10",5,"Sonata","Bronze","—","Re: Corestack","Deovrat Soman","PremKumar B","Jun 14","Awaiting Resolution Confirmation","Unable to reproduce the issue, informed the same to the customer and we are awaiting their response."),
    (220165,"P3","2026-06-10",5,"CoreStack (Internal)","Bronze","—","Trustedtech - HMH - cost for April","Krishna Kumar VJ","Nithin Ram","Jun 10","Awaiting Resolution Confirmation","N/A"),
    (220416,"P2","2026-06-11",4,"Sonata","Bronze","133815","Sonata - Tata Tele CSP processing issue","Deovrat Soman","PremKumar B","Jun 15","Awaiting Resolution Confirmation","Cost has been processed and we are now awaiting customer's confirmation."),
    (220456,"P3","2026-06-11",4,"Kyndryl","Gold","—","US SaaS - Kyndryl Lifelabs - 2 subscriptions are not available","Nagalakshmi N","Avinash Naidu","Jun 15","Awaiting Resolution Confirmation","NA"),
    (221054,"P3","2026-06-12",3,"Kyndryl","Gold","—","US SaaS - Kyndryl - Default dashboards not visible","Nagalakshmi N","Nithin Ram","Jun 12","Awaiting Resolution Confirmation","NA"),
]

OPEN = [
    (219258,"Normal","2026-06-08",7,"CoreStack (Internal)","Bronze","—","Firing: High Priority MSProd App Server Memory Utilisation above 90%","Notify SRE Ops","—","Jun 8","Open",""),
    (219898,"P3","2026-06-10",5,"Cloud Kinetics","Silver","—","Deployment Status Confirmation Required","Service Assurance","PremKumar B","Jun 10","Open",""),
    (220491,"P3","2026-06-11",4,"Kyndryl","Gold","133656","FW: Resources Cost - Beside Tagged and untagged - LifeLabs","Nagalakshmi N","Avinash Naidu","Jun 12","Open",""),
]

ALL_TICKETS = IN_PROGRESS + ON_HOLD + ARC + OPEN

def age_bucket(age):
    if age >= 30: return "30d+"
    if age >= 15: return "15-30d"
    if age >= 8: return "8-14d"
    return "0-7d"

def age_badge_style(age):
    if age >= 30:
        return "background:#FEE2E2;color:#B91C1C"
    if age >= 15:
        return "background:#FEF3C7;color:#B45309"
    if age >= 8:
        return "background:#FEFCE8;color:#713F12"
    return "background:#F0FDF4;color:#14532D"

def priority_badge(p):
    if p == "P2":
        return f'<span style="font-size:10px;font-weight:700;background:#FFF7ED;color:#C2410C;padding:2px 8px;border-radius:4px;">P2</span>'
    elif p == "P3":
        return f'<span style="font-size:10px;font-weight:700;background:#F1F5F9;color:#475569;padding:2px 8px;border-radius:4px;">P3</span>'
    else:
        return f'<span style="font-size:10px;font-weight:700;background:#F1F5F9;color:#475569;padding:2px 8px;border-radius:4px;">{p}</span>'

def status_pill(s):
    if s == "In Progress":
        return '<span style="font-size:10px;font-weight:600;background:#F0FDF4;color:#15803D;padding:2px 7px;border-radius:10px;white-space:nowrap;">In Progress</span>'
    elif s == "On Hold":
        return '<span style="font-size:10px;font-weight:600;background:#FFFBEB;color:#B45309;padding:2px 7px;border-radius:10px;white-space:nowrap;">On Hold</span>'
    elif s == "Awaiting Resolution Confirmation":
        return '<span style="font-size:10px;font-weight:600;background:#EDE9FE;color:#6D28D9;padding:2px 7px;border-radius:10px;white-space:nowrap;">Awaiting Resolution Confirmation</span>'
    else:
        return '<span style="font-size:10px;font-weight:600;background:#EFF6FF;color:#1D4ED8;padding:2px 7px;border-radius:10px;white-space:nowrap;">Open</span>'

def team_badge(ado):
    if ado and ado != "—":
        return '<span style="font-size:9px;font-weight:700;background:#EDE9FE;color:#6D28D9;padding:2px 5px;border-radius:3px;">L2</span>'
    return '<span style="font-size:9px;font-weight:600;background:#F1F5F9;color:#64748B;padding:2px 5px;border-radius:3px;">L1</span>'

def ado_badge(ado):
    if ado and ado != "—":
        return f'<span style="font-size:10px;font-weight:700;background:#F5F3FF;color:#6D28D9;padding:2px 5px;border-radius:4px;">{ado}</span>'
    return ""

def row_bg(ticket_num, status, age):
    # Apply specific background rules
    bg_map = {
        209005: "#FFF0F0",
        211893: "#FFFBF0",
        211895: "#FFFBF0",
        211954: "#FFFBF0",
        217785: "#FFFBF0",
        217515: "#FEFFF0",
        216808: "#FEFFF0",
        216750: "#FEFFF0",
    }
    return bg_map.get(ticket_num, "#FFFFFF")

TABLE_HEADER = '''<tr style="background:#F8FAFC;"><th style="padding:5px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;text-align:left;border-bottom:1px solid #E4E8EF;width:7%;">Ticket</th><th style="padding:5px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;text-align:left;border-bottom:1px solid #E4E8EF;width:22%;">Subject</th><th style="padding:5px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;text-align:left;border-bottom:1px solid #E4E8EF;width:8%;">Priority</th><th style="padding:5px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;text-align:left;border-bottom:1px solid #E4E8EF;width:13%;">Status</th><th style="padding:5px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;text-align:left;border-bottom:1px solid #E4E8EF;width:5%;">Age</th><th style="padding:5px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;text-align:left;border-bottom:1px solid #E4E8EF;width:5%;">Team</th><th style="padding:5px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;text-align:left;border-bottom:1px solid #E4E8EF;width:6%;">ADO #</th><th style="padding:5px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;text-align:left;border-bottom:1px solid #E4E8EF;width:13%;">Raised By</th><th style="padding:5px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;text-align:left;border-bottom:1px solid #E4E8EF;width:13%;">Last Update</th></tr>'''

def ticket_row(t):
    ticket_num, priority, date, age, account, band, ado, subject, raised_by, assignee, last_update, status, reason = t
    bg = row_bg(ticket_num, status, age)
    age_style = age_badge_style(age)
    return f'''<tr style="background:{bg};"><td style="padding:5px 10px;font-size:11px;font-weight:600;color:#2563EB;border-bottom:1px solid #F1F5F9;">#{ticket_num}</td><td style="padding:5px 10px;font-size:11px;color:#334155;border-bottom:1px solid #F1F5F9;">{subject}</td><td style="padding:5px 10px;border-bottom:1px solid #F1F5F9;">{priority_badge(priority)}</td><td style="padding:5px 10px;border-bottom:1px solid #F1F5F9;">{status_pill(status)}</td><td style="padding:5px 10px;border-bottom:1px solid #F1F5F9;"><span style="font-size:10px;font-weight:600;{age_style};padding:2px 6px;border-radius:10px;">{age}d</span></td><td style="padding:5px 10px;border-bottom:1px solid #F1F5F9;">{team_badge(ado)}</td><td style="padding:5px 10px;border-bottom:1px solid #F1F5F9;">{ado_badge(ado)}</td><td style="padding:5px 10px;font-size:10px;color:#475569;border-bottom:1px solid #F1F5F9;">{raised_by}</td><td style="padding:5px 10px;font-size:10px;color:#64748B;border-bottom:1px solid #F1F5F9;">{last_update}</td></tr>'''

def band_badge(band):
    if band == "Platinum":
        return '<span style="font-size:9px;font-weight:700;background:#1E293B;color:#CBD5E1;border:1px solid #334155;border-radius:4px;padding:2px 7px;">👑 Platinum</span>'
    elif band == "Gold":
        return '<span style="font-size:9px;font-weight:700;background:#FFFBEB;color:#92400E;border:1px solid #FDE68A;border-radius:4px;padding:2px 7px;">⭐ Gold</span>'
    else:
        return '<span style="font-size:9px;font-weight:700;background:#F1F5F9;color:#475569;border:1px solid #CBD5E1;border-radius:4px;padding:2px 7px;">🥈 Silver</span>'

def status_pill_small(s):
    if s == "In Progress":
        return '<span style="font-size:10px;background:#F0FDF4;color:#15803D;padding:2px 6px;border-radius:4px;margin-left:4px;">%s</span>'
    elif s == "On Hold":
        return '<span style="font-size:10px;background:#FFFBEB;color:#B45309;padding:2px 6px;border-radius:4px;margin-left:4px;">%s</span>'
    elif s == "Awaiting Resolution Confirmation":
        return '<span style="font-size:10px;background:#EDE9FE;color:#6D28D9;padding:2px 6px;border-radius:4px;margin-left:4px;">%s</span>'
    else:
        return '<span style="font-size:10px;background:#EFF6FF;color:#1D4ED8;padding:2px 6px;border-radius:4px;margin-left:4px;">%s</span>'

def build_account_section(account_name, band, tickets, pills_html):
    rows = "".join(ticket_row(t) for t in tickets)
    return f'''<tr><td style="padding-bottom:14px;"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#fff;border:1px solid #E4E8EF;border-radius:10px;margin-bottom:10px;"><tr><td style="background:#F8FAFC;border-bottom:1px solid #E4E8EF;padding:8px 14px;"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="font-size:13px;font-weight:700;color:#0F172A;">{account_name}&nbsp;{band_badge(band)}</td><td align="right">{pills_html}</td></tr></table></td></tr><tr><td><table width="100%" cellpadding="0" cellspacing="0" border="0">{TABLE_HEADER}{rows}</table></td></tr></table></td></tr>'''

# Build account sections
def pills(*items):
    result = ""
    for label, style_fn in items:
        tmpl = style_fn("")
        # Extract the span tag and insert label
        result += tmpl % label
    return result

def pill_ip(label): return f'<span style="font-size:10px;background:#F0FDF4;color:#15803D;padding:2px 6px;border-radius:4px;margin-left:4px;">{label}</span>'
def pill_hold(label): return f'<span style="font-size:10px;background:#FFFBEB;color:#B45309;padding:2px 6px;border-radius:4px;margin-left:4px;">{label}</span>'
def pill_arc(label): return f'<span style="font-size:10px;background:#EDE9FE;color:#6D28D9;padding:2px 6px;border-radius:4px;margin-left:4px;">{label}</span>'
def pill_open(label): return f'<span style="font-size:10px;background:#EFF6FF;color:#1D4ED8;padding:2px 6px;border-radius:4px;margin-left:4px;">{label}</span>'

# Account data lookup
ticket_map = {t[0]: t for t in ALL_TICKETS}

def get_tickets(*nums):
    return [ticket_map[n] for n in nums]

account_sections = ""

# 1. Neurealm - Platinum
account_sections += build_account_section("Neurealm", "Platinum",
    get_tickets(209005, 211893, 211895),
    pill_hold("3 On Hold"))

# 2. Kyndryl - Gold
account_sections += build_account_section("Kyndryl", "Gold",
    get_tickets(217515, 216808, 220105, 220456, 221054, 220491),
    pill_ip("3 In Progress") + pill_arc("2 Awaiting Resolution Confirmation") + pill_open("1 Open"))

# 3. Logicalis - Gold
account_sections += build_account_section("Logicalis", "Gold",
    get_tickets(205888),
    pill_ip("1 In Progress"))

# 4. Otsuka - Gold
account_sections += build_account_section("Otsuka", "Gold",
    get_tickets(211954, 217606),
    pill_hold("1 On Hold") + pill_arc("1 Awaiting Resolution Confirmation"))

# 5. Synopsys - Gold
account_sections += build_account_section("Synopsys", "Gold",
    get_tickets(220875),
    pill_ip("1 In Progress"))

# 6. Synoptek - Gold
account_sections += build_account_section("Synoptek", "Gold",
    get_tickets(206833),
    pill_ip("1 In Progress"))

# 7. Tata Communications - Gold
account_sections += build_account_section("Tata Communications", "Gold",
    get_tickets(211192, 216750),
    pill_ip("2 In Progress"))

# 8. Cloud Kinetics - Silver
account_sections += build_account_section("Cloud Kinetics", "Silver",
    get_tickets(217785, 219898),
    pill_hold("1 On Hold") + pill_open("1 Open"))

# 9. Getronics - Silver
account_sections += build_account_section("Getronics", "Silver",
    get_tickets(219361),
    pill_arc("1 Awaiting Resolution Confirmation"))

# ──────────────────────────────────────────────
# HTML GENERATION
# ──────────────────────────────────────────────

html = f'''<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"><title>Daily Incident Report · June 15, 2026</title></head>
<body style="margin:0;padding:0;background:#F4F6F9;font-family:Arial,sans-serif;font-size:13px;color:#1A2035;">
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#F4F6F9;"><tr><td align="center" style="padding:20px 16px 40px;">
<table width="980" cellpadding="0" cellspacing="0" border="0" style="max-width:980px;width:100%;">
<tr><td style="background:#fff;border:1px solid #E4E8EF;border-radius:10px;padding:18px 28px 16px;">
  <div style="font-size:17px;font-weight:700;color:#0F172A;">Daily Incident Report</div>
  <div style="font-size:11px;color:#64748B;margin-top:3px;"><b style="color:#334155;">Period:</b>&nbsp;June 15, 2026 · 19:30 IST&nbsp;&nbsp;<b style="color:#334155;">Dept:</b>&nbsp;CoreStack Support</div>
</td></tr><tr><td height="14"></td></tr>
<tr><td style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:#94A3B8;padding-bottom:10px;">Summary</td></tr>
<tr><td style="padding-bottom:14px;"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr>
  <td width="20%" style="padding-right:10px;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#fff;border:1px solid #E4E8EF;border-radius:10px;border-top:3px solid #3B82F6;"><tr><td style="padding:12px 14px;"><div style="font-size:10px;font-weight:600;color:#64748B;text-transform:uppercase;">New</div><div style="font-size:26px;font-weight:700;color:#3B82F6;line-height:1.1;margin:4px 0 2px;">3</div><div style="font-size:10px;color:#94A3B8;">opened</div></td></tr></table></td>
  <td width="20%" style="padding-right:10px;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#fff;border:1px solid #E4E8EF;border-radius:10px;border-top:3px solid #10B981;"><tr><td style="padding:12px 14px;"><div style="font-size:10px;font-weight:600;color:#64748B;text-transform:uppercase;">In Progress</div><div style="font-size:26px;font-weight:700;color:#10B981;line-height:1.1;margin:4px 0 2px;">15</div><div style="font-size:10px;color:#94A3B8;">being worked</div></td></tr></table></td>
  <td width="20%" style="padding-right:10px;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#fff;border:1px solid #E4E8EF;border-radius:10px;border-top:3px solid #F59E0B;"><tr><td style="padding:12px 14px;"><div style="font-size:10px;font-weight:600;color:#64748B;text-transform:uppercase;">On Hold</div><div style="font-size:26px;font-weight:700;color:#F59E0B;line-height:1.1;margin:4px 0 2px;">11</div><div style="font-size:10px;color:#94A3B8;">pending / monitoring</div></td></tr></table></td>
  <td width="20%" style="padding-right:10px;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#fff;border:1px solid #E4E8EF;border-radius:10px;border-top:3px solid #8B5CF6;"><tr><td style="padding:12px 14px;"><div style="font-size:10px;font-weight:600;color:#64748B;text-transform:uppercase;">Awaiting Confirmation</div><div style="font-size:26px;font-weight:700;color:#8B5CF6;line-height:1.1;margin:4px 0 2px;">11</div><div style="font-size:10px;color:#94A3B8;">awaiting customer</div></td></tr></table></td>
  <td width="20%" style="padding-right:0px;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#fff;border:1px solid #E4E8EF;border-radius:10px;border-top:3px solid #EF4444;"><tr><td style="padding:12px 14px;"><div style="font-size:10px;font-weight:600;color:#64748B;text-transform:uppercase;">With L2 (PGS)</div><div style="font-size:26px;font-weight:700;color:#EF4444;line-height:1.1;margin:4px 0 2px;">14</div><div style="font-size:10px;color:#94A3B8;">ADO linked</div></td></tr></table></td>
</tr></table></td></tr>
<tr><td style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:#94A3B8;padding-bottom:10px;">Open Ticket Aging</td></tr>
<tr><td style="padding-bottom:14px;"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr>
  <td width="25%" style="padding-right:10px;"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#FEF2F2;border:1px solid #FECACA;border-radius:8px;"><tr><td style="padding:10px 14px;"><table width="100%" cellpadding="0" cellspacing="0"><tr><td style="font-size:11px;font-weight:600;color:#991B1B;">30d+</td><td align="right" style="font-size:26px;font-weight:700;color:#DC2626;">3</td></tr></table></td></tr></table></td>
  <td width="25%" style="padding-right:10px;"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#FFFBEB;border:1px solid #FDE68A;border-radius:8px;"><tr><td style="padding:10px 14px;"><table width="100%" cellpadding="0" cellspacing="0"><tr><td style="font-size:11px;font-weight:600;color:#92400E;">15-30d</td><td align="right" style="font-size:26px;font-weight:700;color:#D97706;">6</td></tr></table></td></tr></table></td>
  <td width="25%" style="padding-right:10px;"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#FEFCE8;border:1px solid #FEF08A;border-radius:8px;"><tr><td style="padding:10px 14px;"><table width="100%" cellpadding="0" cellspacing="0"><tr><td style="font-size:11px;font-weight:600;color:#713F12;">8-14d</td><td align="right" style="font-size:26px;font-weight:700;color:#CA8A04;">5</td></tr></table></td></tr></table></td>
  <td width="25%"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#F0FDF4;border:1px solid #BBF7D0;border-radius:8px;"><tr><td style="padding:10px 14px;"><table width="100%" cellpadding="0" cellspacing="0"><tr><td style="font-size:11px;font-weight:600;color:#14532D;">0-7d</td><td align="right" style="font-size:26px;font-weight:700;color:#16A34A;">26</td></tr></table></td></tr></table></td>
</tr></table></td></tr>
<tr><td style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:#94A3B8;padding-bottom:10px;">Tickets by Account Band</td></tr>
<tr><td style="padding-bottom:14px;"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr>
  <td width="33%" style="padding-right:10px;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#1E293B;border:1px solid #334155;border-radius:10px;"><tr><td style="padding:14px 16px;"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="font-size:11px;font-weight:700;color:#CBD5E1;text-transform:uppercase;">👑 Platinum</td><td align="right" style="font-size:28px;font-weight:700;color:#CBD5E1;">3</td></tr></table></td></tr></table></td>
  <td width="33%" style="padding-right:10px;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#451A03;border:1px solid #78350F;border-radius:10px;"><tr><td style="padding:14px 16px;"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="font-size:11px;font-weight:700;color:#FBBF24;text-transform:uppercase;">⭐ Gold</td><td align="right" style="font-size:28px;font-weight:700;color:#FBBF24;">13</td></tr></table></td></tr></table></td>
  <td width="33%" style="padding-right:10px;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#1E3A5F;border:1px solid #1E4976;border-radius:10px;"><tr><td style="padding:14px 16px;"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="font-size:11px;font-weight:700;color:#93C5FD;text-transform:uppercase;">🥈 Silver</td><td align="right" style="font-size:28px;font-weight:700;color:#93C5FD;">3</td></tr></table></td></tr></table></td>
</tr></table></td></tr>
<tr><td style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:#94A3B8;padding-bottom:10px;">Ticket Detail by Account (Platinum · Gold · Silver)</td></tr>
<tr><td style="padding-bottom:14px;"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#F8FAFC;border:1px solid #E4E8EF;border-radius:10px;"><tr><td style="padding:14px 18px;">
  <div style="font-size:10px;font-weight:700;color:#64748B;text-transform:uppercase;letter-spacing:0.06em;margin-bottom:10px;">Reference</div>
  <table width="100%" cellpadding="0" cellspacing="0" border="0"><tr>
    <td width="50%" valign="top" style="padding-right:24px;border-right:1px solid #E4E8EF;">
      <div style="font-size:10px;font-weight:700;color:#B45309;margin-bottom:4px;">🟡 On Hold</div>
      <div style="font-size:10px;color:#64748B;line-height:1.7;">Ticket kept for monitoring, backfill in progress, pending information from the customer or third parties like Azure / AWS / GCP / OCI.</div>
    </td>
    <td width="50%" valign="top" style="padding-left:24px;">
      <div style="font-size:10px;font-weight:700;color:#6D28D9;margin-bottom:4px;">🟣 Awaiting Resolution Confirmation</div>
      <div style="font-size:10px;color:#64748B;line-height:1.7;">Issue fixed / clarification has been communicated. Awaiting resolution confirmation from the customer for closure.</div>
    </td>
  </tr></table>
</td></tr></table></td></tr>

{account_sections}

<tr><td><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#F0F9FF;border:1px solid #BAE6FD;border-radius:10px;">
  <tr><td style="padding:14px 20px;"><span style="font-size:11px;color:#64748B;">📊 Full ticket dump (Excel): </span>
  <a href="CS_Daily_Incident_Report_20260615.xlsx" style="font-size:11px;font-weight:600;color:#2563EB;text-decoration:none;">Download Daily Incident Report Excel →</a>
  </td></tr></table></td></tr>
</table></td></tr></table></body></html>
'''

with open("/home/user/GD/CS_Daily_Incident_Report_20260615.html", "w", encoding="utf-8") as f:
    f.write(html)

print("HTML written.")

# ──────────────────────────────────────────────
# EXCEL GENERATION
# ──────────────────────────────────────────────

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

def make_fill(hex_color):
    return PatternFill(start_color=hex_color.lstrip("#"), end_color=hex_color.lstrip("#"), fill_type="solid")

def make_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color.lstrip("#"), size=size)

def make_border():
    thin = Side(style="thin", color="D1D5DB")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def make_align(wrap=True, h="left", v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def apply_header(ws, headers, header_fill, header_font):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = make_align(h="center")
        cell.border = make_border()
    ws.row_dimensions[1].height = 20

def band_row_fill(band):
    fills = {
        "Platinum": ("CBD5E1", "1E293B"),
        "Gold": ("FEF9C3", "92400E"),
        "Silver": ("DBEAFE", "1E4976"),
        "Bronze": ("FFF7ED", "9A3412"),
    }
    return fills.get(band, ("FFFFFF", "000000"))

def apply_data_row(ws, row_num, values, band):
    fill_hex, text_hex = band_row_fill(band)
    ws.append(values)
    r = ws.row_dimensions[row_num]
    r.height = 30
    for cell in ws[row_num]:
        cell.fill = make_fill("#" + fill_hex)
        cell.font = make_font(color="#" + text_hex)
        cell.alignment = make_align()
        cell.border = make_border()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

COLS_14 = ["Ticket#", "Subject", "Priority", "Account", "Band", "Assignee", "Status", "Age(days)", "Age Bucket", "L1/L2", "ADO Link", "Raised By", "Last Update", "Hold/ARC Reason"]
COLS_13 = ["Ticket#", "Subject", "Priority", "Account", "Band", "Assignee", "Status", "Age(days)", "Age Bucket", "L1/L2", "ADO Link", "Raised By", "Last Update"]
WIDTHS_14 = [10, 18, 10, 40, 8, 26, 8, 10, 6, 10, 16, 36, 16, 12]
WIDTHS_13 = [10, 18, 10, 44, 8, 26, 8, 10, 6, 10, 16, 16, 12]

def ticket_to_14(t):
    ticket_num, priority, date, age, account, band, ado, subject, raised_by, assignee, last_update, status, reason = t
    l1l2 = "L2" if ado and ado != "—" else "L1"
    ado_val = ado if ado and ado != "—" else ""
    return [ticket_num, subject, priority, account, band, assignee, status, age, age_bucket(age), l1l2, ado_val, raised_by, last_update, reason]

def ticket_to_13(t):
    row = ticket_to_14(t)
    return row[:13]

# ── Sheet 1: Summary ──
ws = wb.create_sheet("Summary")
hdr_fill = make_fill("#1E293B")
hdr_font = make_font(bold=True, color="#FFFFFF", size=10)
apply_header(ws, ["Status", "Count", "%"], hdr_fill, hdr_font)
summary_data = [
    ("Open", 3, "7.5%"),
    ("In Progress", 15, "37.5%"),
    ("On Hold", 11, "27.5%"),
    ("Awaiting Resolution Confirmation", 11, "27.5%"),
    ("Total", 40, "100%"),
]
for i, row in enumerate(summary_data, 2):
    ws.append(list(row))
    ws.row_dimensions[i].height = 30
    for cell in ws[i]:
        cell.alignment = make_align()
        cell.border = make_border()
        if row[0] == "Total":
            cell.font = make_font(bold=True)
ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 10

# ── Sheet 2: All Tickets ──
ws = wb.create_sheet("All Tickets")
apply_header(ws, COLS_14, make_fill("#1E293B"), make_font(bold=True, color="#FFFFFF"))
set_col_widths(ws, WIDTHS_14)
for i, t in enumerate(ALL_TICKETS, 2):
    apply_data_row(ws, i, ticket_to_14(t), t[5])

# ── Sheet 3: Platinum Gold Silver ──
ws = wb.create_sheet("Platinum Gold Silver")
apply_header(ws, COLS_14, make_fill("#334155"), make_font(bold=True, color="#FFFFFF"))
set_col_widths(ws, WIDTHS_14)
pgs_tickets = [t for t in ALL_TICKETS if t[5] in ("Platinum", "Gold", "Silver")]
for i, t in enumerate(pgs_tickets, 2):
    apply_data_row(ws, i, ticket_to_14(t), t[5])

# ── Sheet 4: New Open ──
ws = wb.create_sheet("New Open")
apply_header(ws, COLS_13, make_fill("#1D4ED8"), make_font(bold=True, color="#FFFFFF"))
set_col_widths(ws, WIDTHS_13)
for i, t in enumerate(OPEN, 2):
    apply_data_row(ws, i, ticket_to_13(t), t[5])

# ── Sheet 5: In Progress ──
ws = wb.create_sheet("In Progress")
apply_header(ws, COLS_13, make_fill("#15803D"), make_font(bold=True, color="#FFFFFF"))
set_col_widths(ws, WIDTHS_13)
for i, t in enumerate(IN_PROGRESS, 2):
    apply_data_row(ws, i, ticket_to_13(t), t[5])

# ── Sheet 6: On Hold ──
ws = wb.create_sheet("On Hold")
apply_header(ws, COLS_14, make_fill("#B45309"), make_font(bold=True, color="#FFFFFF"))
set_col_widths(ws, WIDTHS_14)
for i, t in enumerate(ON_HOLD, 2):
    apply_data_row(ws, i, ticket_to_14(t), t[5])

# ── Sheet 7: Awaiting Resolution ──
ws = wb.create_sheet("Awaiting Resolution")
apply_header(ws, COLS_14, make_fill("#6D28D9"), make_font(bold=True, color="#FFFFFF"))
set_col_widths(ws, WIDTHS_14)
for i, t in enumerate(ARC, 2):
    apply_data_row(ws, i, ticket_to_14(t), t[5])

# ── Sheet 8: L2 Tickets ──
ws = wb.create_sheet("L2 Tickets")
apply_header(ws, COLS_14, make_fill("#6D28D9"), make_font(bold=True, color="#FFFFFF"))
set_col_widths(ws, WIDTHS_14)
l2_tickets = [t for t in ALL_TICKETS if t[6] and t[6] != "—"]
for i, t in enumerate(l2_tickets, 2):
    apply_data_row(ws, i, ticket_to_14(t), t[5])

# ── Sheet 9: Aging View ──
ws = wb.create_sheet("Aging View")
AGING_COLS = ["Bucket", "Ticket#", "Subject", "Priority", "Account", "Band", "Assignee", "Status"]
apply_header(ws, AGING_COLS, make_fill("#1E293B"), make_font(bold=True, color="#FFFFFF"))
ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 40
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 28
ws.column_dimensions["F"].width = 10
ws.column_dimensions["G"].width = 22
ws.column_dimensions["H"].width = 20

buckets = [
    ("30d+", 3),
    ("15-30d", 6),
    ("8-14d", 5),
    ("0-7d", 26),
]

bucket_tickets = {
    "30d+": sorted([t for t in ALL_TICKETS if t[3] >= 30], key=lambda x: -x[3]),
    "15-30d": sorted([t for t in ALL_TICKETS if 15 <= t[3] < 30], key=lambda x: -x[3]),
    "8-14d": sorted([t for t in ALL_TICKETS if 8 <= t[3] < 15], key=lambda x: -x[3]),
    "0-7d": sorted([t for t in ALL_TICKETS if t[3] < 8], key=lambda x: -x[3]),
}

current_row = 2
for bucket, count in buckets:
    # Header row for bucket
    header_label = f"{bucket} ({count} tickets)"
    ws.append([header_label, "", "", "", "", "", "", ""])
    ws.row_dimensions[current_row].height = 20
    bucket_fill = {"30d+": "#FEE2E2", "15-30d": "#FEF3C7", "8-14d": "#FEFCE8", "0-7d": "#F0FDF4"}[bucket]
    bucket_txt = {"30d+": "#B91C1C", "15-30d": "#B45309", "8-14d": "#713F12", "0-7d": "#14532D"}[bucket]
    for cell in ws[current_row]:
        cell.fill = make_fill(bucket_fill)
        cell.font = make_font(bold=True, color=bucket_txt)
        cell.alignment = make_align(h="left")
        cell.border = make_border()
    current_row += 1

    for t in bucket_tickets[bucket]:
        ticket_num, priority, date, age, account, band, ado, subject, raised_by, assignee, last_update, status, reason = t
        row_vals = [bucket, ticket_num, subject, priority, account, band, assignee, status]
        ws.append(row_vals)
        ws.row_dimensions[current_row].height = 30
        fill_hex, text_hex = band_row_fill(band)
        for cell in ws[current_row]:
            cell.fill = make_fill("#" + fill_hex)
            cell.font = make_font(color="#" + text_hex)
            cell.alignment = make_align()
            cell.border = make_border()
        current_row += 1

wb.save("/home/user/GD/CS_Daily_Incident_Report_20260615.xlsx")
print("Excel written.")
print("Done!")
