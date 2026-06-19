"""
report_generator.py
Generates the Daily Incident Report HTML and Excel from live Zoho ticket data.
"""
from __future__ import annotations

import io
import re
from datetime import date, datetime, timezone, timedelta

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── IST offset ────────────────────────────────────────────────────────────────
IST = timedelta(hours=5, minutes=30)


# ── Band / display-name classification ───────────────────────────────────────
# Band classifications from Customer Adoption Report (Apr 2026)
ACCOUNT_BAND: dict[str, str] = {
    # Platinum
    "neurealm":           "Platinum",
    "taylor farms":       "Platinum",
    "taylor_farms":       "Platinum",
    "taylorfarms":        "Platinum",
    "taylor farm":        "Platinum",
    # Platinum — named strategic accounts
    "blackstone":         "Platinum",
    # Gold
    "otsuka":             "Gold",
    "otsuka-us":          "Gold",
    "kyndryl":            "Gold",
    "tatacommunications": "Gold",
    "tata communications":"Gold",
    "au.logicalis":       "Gold",
    "logicalis":          "Gold",
    "synopsys":           "Gold",
    "synoptek":           "Gold",
    "bluemantis":         "Gold",
    "blue mantis":        "Gold",
    "hitachi":            "Gold",
    "tcl":                "Gold",
    "it1":                "Gold",
    "gevernova":          "Gold",
    "ge vernova":         "Gold",
    # Silver
    "getronics":          "Silver",
    "cloud-kinetics":     "Silver",
    "cloud kinetics":     "Silver",
    "cloudelligent":      "Silver",
    "aliando":            "Silver",
    "virtusa":            "Silver",
    "sonata":             "Silver",
    "sonata-software":    "Silver",
    "sonata software":    "Silver",
    "damac":              "Silver",
    "nbf":                "Silver",
    "ntt":                "Silver",
    "microland":          "Silver",
}

DISPLAY_NAMES: dict[str, str] = {
    "neurealm":           "Neurealm",
    "taylor farms":       "Taylor Farms",
    "taylor_farms":       "Taylor Farms",
    "taylorfarms":        "Taylor Farms",
    "taylor farm":        "Taylor Farms",
    "otsuka":             "Otsuka",
    "otsuka-us":          "Otsuka",
    "kyndryl":            "Kyndryl",
    "tatacommunications": "Tata Communications",
    "tata communications":"Tata Communications",
    "au.logicalis":       "Logicalis",
    "logicalis":          "Logicalis",
    "synopsys":           "Synopsys",
    "synoptek":           "Synoptek",
    "bluemantis":         "BlueMantis",
    "blue mantis":        "BlueMantis",
    "hitachi":            "Hitachi",
    "tcl":                "TCL",
    "it1":                "IT1",
    "gevernova":          "GE Vernova",
    "ge vernova":         "GE Vernova",
    "getronics":          "Getronics",
    "cloud-kinetics":     "Cloud Kinetics",
    "cloud kinetics":     "Cloud Kinetics",
    "cloudelligent":      "Cloudelligent",
    "aliando":            "Aliando",
    "virtusa":            "Virtusa",
    "sonata":             "Sonata",
    "sonata-software":    "Sonata",
    "sonata software":    "Sonata",
    "damac":              "Damac",
    "nbf":                "NBF",
    "ntt":                "NTT",
    "microland":          "Microland",
    "ltts":               "LTTS",
    "core42":             "Core42",
    "blackstone":         "Blackstone",
}

BLACKSTONE_KEYWORDS = []  # now handled via ACCOUNT_BAND as Platinum


def _norm(s: str) -> str:
    return s.strip().lower() if s else ""


def get_band(account_name: str) -> str:
    key = _norm(account_name)
    if not key:
        return "Bronze"
    for k, band in ACCOUNT_BAND.items():
        if k in key or key in k:
            return band
    return "Bronze"


def get_display(account_name: str) -> str:
    key = _norm(account_name)
    for k, disp in DISPLAY_NAMES.items():
        if k == key:
            return disp
        if k in key:
            return disp
    if any(kw in key for kw in BLACKSTONE_KEYWORDS):
        return "Blackstone"
    return account_name or "Unknown"


# ── Zoho ticket → internal tuple ─────────────────────────────────────────────
def parse_ticket(raw: dict, today: date) -> dict:
    """
    Returns a normalized ticket dict with keys:
      num, subject, priority, account_raw, band, display, ado,
      created_date, last_updated, contact, assignee, status, reason, age, bucket
    """
    num = raw.get("ticketNumber", "")
    subject = raw.get("subject", "")
    _pri_map = {"Critical": "P1", "High": "P2", "Medium": "P3", "Normal": "P3", "Low": "P4"}
    _raw_pri = raw.get("priority") or "Normal"
    priority = _pri_map.get(_raw_pri, _raw_pri if _raw_pri.startswith("P") else "P3")
    status = raw.get("status", "")

    cf = raw.get("cf") or {}
    custom_fields = raw.get("customFields") or {}

    # Account name — live API returns account=null; name is in cf_customer
    account_raw = ""
    if raw.get("account"):
        account_raw = raw["account"].get("accountName") or raw["account"].get("name") or ""
    if not account_raw:
        account_raw = raw.get("accountName", "")
    # Try cf_customer → cf_pov → email domain (in order)
    _IGNORE_ACCOUNTS = {"internal", "corestack", "corestack internal", "unknown", "n/a", "na", ""}
    if not account_raw or _norm(account_raw) in _IGNORE_ACCOUNTS:
        account_raw = cf.get("cf_customer") or custom_fields.get("Customer") or ""
    if not account_raw or _norm(account_raw) in _IGNORE_ACCOUNTS:
        account_raw = cf.get("cf_pov") or custom_fields.get("POV") or ""
    if _norm(account_raw) in _IGNORE_ACCOUNTS:
        account_raw = ""

    if not account_raw:
        email = raw.get("email") or ""
        domain = email.split("@")[-1].split(".")[0] if "@" in email else ""
        _domain_map = {
            "neurealm": "Neurealm", "otsuka": "Otsuka", "otsuka-us": "Otsuka",
            "kyndryl": "Kyndryl", "synopsys": "Synopsys", "synoptek": "Synoptek",
            "getronics": "Getronics", "logicalis": "Logicalis",
            "tata": "Tata Communications", "synoptek": "Synoptek",
        }
        account_raw = _domain_map.get(domain.lower(), "")

    # ADO — cf_ado_reference holds a full Azure DevOps URL; extract the work item number
    ado = ""
    ado_raw = (cf.get("cf_ado_reference") or custom_fields.get("ADO reference") or "").strip()
    if ado_raw and ado_raw.upper() not in ("NA", "N/A", ""):
        m = re.search(r"/(\d+)\s*$", ado_raw)
        ado = m.group(1) if m else ado_raw
    if not ado:
        for key in ("cf_ado_number", "cf_ado", "cf_adoNumber", "cf_ado_link", "cf_azure_devops"):
            if cf.get(key):
                ado = str(cf[key]).strip()
                break

    # Dates — Zoho returns ISO strings with timezone
    def _parse_dt(s: str | None) -> date | None:
        if not s:
            return None
        try:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
            return (dt + IST).date()
        except Exception:
            return None

    created_date = _parse_dt(raw.get("createdTime")) or today
    last_updated_date = _parse_dt(raw.get("modifiedTime")) or today

    def _fmt_date(d: date) -> str:
        months = ["Jan","Feb","Mar","Apr","May","Jun",
                  "Jul","Aug","Sep","Oct","Nov","Dec"]
        return f"{months[d.month-1]} {d.day:02d}"

    # Contact — detail API returns contacts as list (include=contacts) or single object
    contact = ""
    contacts_list = raw.get("contacts") or []
    if contacts_list and isinstance(contacts_list, list):
        c = contacts_list[0]
        contact = ((c.get("firstName") or "") + " " + (c.get("lastName") or "")).strip()
    if not contact and raw.get("contact"):
        c = raw["contact"]
        contact = ((c.get("firstName") or "") + " " + (c.get("lastName") or "")).strip()
    if not contact:
        contact = raw.get("contactName") or raw.get("requesterName") or ""
    # Strip email fallback — show name only, not email address
    if not contact or "@" in contact:
        contact = raw.get("contactName") or raw.get("requesterName") or ""

    # Assignee — map known IDs to names
    _assignee_map = {
        "100599000004409465": "Ganga Reddy",
        "100599000004409501": "PremKumar B",
        "100599000004409537": "Nithin Ram",
        "100599000004409573": "Avinash Naidu",
        "100599000049929001": "Deepesh H",
        "100599000000176484": "Aadhithya S",
        "100599000000176268": "Gnanadesigan A",
        "100599000004648021": "Logesh S",
    }
    assignee = ""
    if raw.get("assignee"):
        a = raw["assignee"]
        assignee = ((a.get("firstName") or "") + " " + (a.get("lastName") or "")).strip()
    if not assignee:
        assignee = raw.get("assigneeName") or ""
    if not assignee:
        aid = str(raw.get("assigneeId") or "")
        assignee = _assignee_map.get(aid, "")

    # Reason — on-hold/awaiting reason from custom field
    reason = (cf.get("cf_on_hold_or_awaiting_confirmation_reason")
              or custom_fields.get("On Hold or Awaiting Confirmation Reason")
              or raw.get("resolution") or raw.get("reasonForOnHold") or "")

    age_days = (today - created_date).days
    bkt = _bucket(age_days)

    # If account still empty, try to detect known customer names in subject line
    if not account_raw:
        _subj_lower = _norm(subject)
        _SUBJ_KEYWORDS = [
            "taylor farms", "taylor_farms", "blackstone", "neurealm",
            "otsuka", "kyndryl", "logicalis", "synopsys", "synoptek",
            "bluemantis", "hitachi", "getronics", "virtusa", "microland",
            "tata communications", "ge vernova",
        ]
        for kw in _SUBJ_KEYWORDS:
            if kw in _subj_lower:
                account_raw = kw
                break

    band = get_band(account_raw)
    display = get_display(account_raw)

    return {
        "num": str(num),
        "subject": subject,
        "priority": priority,
        "account_raw": account_raw,
        "band": band,
        "display": display,
        "ado": ado,
        "created_date": created_date,
        "created_str": created_date.isoformat(),
        "last_updated": _fmt_date(last_updated_date),
        "contact": contact,
        "assignee": assignee,
        "status": status,
        "reason": reason,
        "age": age_days,
        "bucket": bkt,
        "team": "L2" if ado else "L1",
    }


def _bucket(a: int) -> str:
    if a <= 7:   return "0-7d"
    if a <= 14:  return "8-14d"
    if a <= 30:  return "15-30d"
    return "30d+"


# ── Style helpers ─────────────────────────────────────────────────────────────
def _f(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)


def _font(hex6: str, bold=False, sz=10) -> Font:
    return Font(color=hex6, bold=bold, size=sz)


AL_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

AGE_STYLES    = {"0-7d":  ("F0FDF4","14532D"), "8-14d": ("FEFCE8","713F12"),
                 "15-30d":("FEF3C7","B45309"),  "30d+":  ("FEE2E2","B91C1C")}
PRI_STYLES    = {"P2": ("FFF7ED","C2410C"), "P3": ("F1F5F9","475569")}
STATUS_STYLES = {"Open":("DBEAFE","1D4ED8"), "In Progress":("DCFCE7","15803D"),
                 "On Hold":("FEF9C3","B45309"),
                 "Awaiting Resolution Confirmation":("EDE9FE","6D28D9")}
BAND_STYLES   = {"Platinum":("CBD5E1","1E293B"), "Gold":("FEF9C3","92400E"),
                 "Silver":("DBEAFE","1E4976"),   "Bronze":("FFF7ED","9A3412")}
BAND_BADGE    = {"Platinum":"Platinum","Gold":"Gold","Silver":"Silver",
                 "Bronze":"Bronze","Blackstone-Bronze":"Bronze"}

ROW_BG = {
    ("On Hold","30d+"):      "FFF0F0",
    ("On Hold","15-30d"):    "FFFBF0",
    ("On Hold","8-14d"):     "FFFBF0",
    ("In Progress","8-14d"): "FEFFF0",
}


def _row_bg(status: str, bkt: str) -> str:
    return ROW_BG.get((status, bkt), "FFFFFF")


# ── Excel helpers ─────────────────────────────────────────────────────────────
W14 = [10, 18, 10, 40, 8, 26, 8, 10, 6, 10, 16, 36, 16, 12]
W13 = [10, 18, 10, 44, 8, 26, 8, 10, 6, 10, 16, 16, 12]
H14 = ["Ticket #","Customer","Band","Subject","Priority","Status","Age","Bucket",
       "Team","ADO #","Raised By","Reason","Last Updated","Created"]
H13 = ["Ticket #","Customer","Band","Subject","Priority","Status","Age","Bucket",
       "Team","ADO #","Raised By","Last Updated","Created"]


def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_header(ws, row_num: int, headers: list[str], fill_hex: str):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row_num, i, h)
        c.fill = _f(fill_hex)
        c.font = Font(color="FFFFFF", bold=True, size=10)
        c.alignment = AL_CENTER
    ws.row_dimensions[row_num].height = 25


def _write_row(ws, row_num: int, t: dict, with_reason: bool = True):
    bkt = t["bucket"]
    status = t["status"]
    rbg = _row_bg(status, bkt)
    band = t["band"]
    af, afont = AGE_STYLES[bkt]
    pf, pfont = PRI_STYLES.get(t["priority"], ("F1F5F9","475569"))
    sf, sfont = STATUS_STYLES.get(status, ("FFFFFF","000000"))
    bf, bfont = BAND_STYLES.get(band, ("FFF7ED","9A3412"))
    badge_label = BAND_BADGE.get(band, "Bronze")

    cells: list[tuple] = [
        (f"#{t['num']}",    rbg,      "2563EB", True),   # A Ticket #
        (t["display"],      rbg,      "0F172A", False),  # B Customer
        (badge_label,       bf,       bfont,    True),   # C Band
        (t["subject"],      rbg,      "0F172A", False),  # D Subject
        (t["priority"],     pf,       pfont,    True),   # E Priority
        (status,            sf,       sfont,    True),   # F Status
        (f"{t['age']}d",    af,       afont,    True),   # G Age
        (bkt,               af,       afont,    False),  # H Bucket
        (t["team"],         "EDE9FE" if t["team"]=="L2" else "F1F5F9",
                            "6D28D9" if t["team"]=="L2" else "64748B",
                            t["team"]=="L2"),             # I Team
        (t["ado"],          "F5F3FF" if t["ado"] else "FFFFFF",
                            "6D28D9", bool(t["ado"])),   # J ADO #
        (t["contact"],      "F8FAFC", "334155", False),  # K Raised By
    ]

    if with_reason:
        cells.append((t["reason"],       "FFFBEB", "92400E", False))  # L Reason
        cells.append((t["last_updated"], rbg,      "0F172A", False))  # M Last Updated
        cells.append((_created_disp(t["created_date"]), rbg, "0F172A", False))  # N Created
    else:
        cells.append((t["last_updated"], rbg,      "0F172A", False))  # L Last Updated
        cells.append((_created_disp(t["created_date"]), rbg, "0F172A", False))  # M Created

    for col_idx, (val, fill_hex, font_hex, bold) in enumerate(cells, 1):
        c = ws.cell(row_num, col_idx, val)
        c.fill = _f(fill_hex)
        c.font = Font(color=font_hex, bold=bold, size=9)
        c.alignment = AL_LEFT
    ws.row_dimensions[row_num].height = 30


def _created_disp(d: date) -> str:
    months = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]
    return f"{months[d.month-1]} {d.day:02d}"


# ── Public API ────────────────────────────────────────────────────────────────
def generate_excel(tickets: list[dict], today: date) -> bytes:
    """Return Excel workbook bytes from normalized ticket dicts."""
    mo = ["January","February","March","April","May","June",
          "July","August","September","October","November","December"]
    title_date = f"{mo[today.month-1]} {today.day}, {today.year}"

    wb = Workbook()

    # -- Summary sheet
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 10
    ws1.column_dimensions["C"].width = 12

    ws1["A1"] = f"Daily Incident Report — {title_date}"
    ws1["A1"].font = Font(color="0F172A", bold=True, size=12)
    ws1["A2"] = f"Period: {title_date} · 18:00 IST"
    ws1["A2"].font = Font(color="64748B", size=10)
    ws1.row_dimensions[1].height = 22
    ws1.row_dimensions[2].height = 18
    ws1.row_dimensions[3].height = 8

    _write_header(ws1, 4, ["Status","Count","% of Total"], "1E293B")

    by_status: dict[str, list] = {}
    for t in tickets:
        by_status.setdefault(t["status"], []).append(t)

    total = len(tickets)
    open_ = by_status.get("Open", [])
    ip    = by_status.get("In Progress", [])
    oh    = by_status.get("On Hold", [])
    arc   = by_status.get("Awaiting Resolution Confirmation", [])

    for i, (s, lst) in enumerate([
        ("Open", open_), ("In Progress", ip),
        ("On Hold", oh), ("Awaiting Resolution Confirmation", arc),
        ("Total", tickets)
    ], 5):
        n = len(lst)
        ws1[f"A{i}"] = s
        ws1[f"B{i}"] = n
        ws1[f"C{i}"] = f"{n/total*100:.1f}%" if s != "Total" else "100%"
        ws1.row_dimensions[i].height = 22
        if s != "Total":
            sf, sfont = STATUS_STYLES.get(s, ("FFFFFF","000000"))
            for col in ["A","B","C"]:
                c = ws1[f"{col}{i}"]
                c.fill = _f(sf)
                c.font = Font(color=sfont, bold=True, size=10)
        else:
            for col in ["A","B","C"]:
                ws1[f"{col}{i}"].font = Font(bold=True, size=10)

    # -- All Tickets
    ws2 = wb.create_sheet("All Tickets")
    _set_widths(ws2, W14)
    _write_header(ws2, 1, H14, "1E293B")
    for r, t in enumerate(tickets, 2):
        _write_row(ws2, r, t, with_reason=True)

    # -- Platinum Gold Silver
    ws3 = wb.create_sheet("Platinum Gold Silver")
    _set_widths(ws3, W14)
    _write_header(ws3, 1, H14, "334155")
    r = 2
    for t in tickets:
        if t["band"] in ("Platinum","Gold","Silver"):
            _write_row(ws3, r, t, with_reason=True); r += 1

    # -- New
    ws4 = wb.create_sheet("New")
    _set_widths(ws4, W13)
    _write_header(ws4, 1, H13, "1D4ED8")
    for r, t in enumerate(open_, 2):
        _write_row(ws4, r, t, with_reason=False)

    # -- In Progress
    ws5 = wb.create_sheet("In Progress")
    _set_widths(ws5, W13)
    _write_header(ws5, 1, H13, "15803D")
    for r, t in enumerate(ip, 2):
        _write_row(ws5, r, t, with_reason=False)

    # -- On Hold
    ws6 = wb.create_sheet("On Hold")
    _set_widths(ws6, W14)
    _write_header(ws6, 1, H14, "B45309")
    for r, t in enumerate(oh, 2):
        _write_row(ws6, r, t, with_reason=True)

    # -- Awaiting Resolution
    ws7 = wb.create_sheet("Awaiting Resolution")
    _set_widths(ws7, W14)
    _write_header(ws7, 1, H14, "6D28D9")
    for r, t in enumerate(arc, 2):
        _write_row(ws7, r, t, with_reason=True)

    # -- L2 Tickets
    ws8 = wb.create_sheet("L2 Tickets")
    _set_widths(ws8, W13)
    _write_header(ws8, 1, H13, "6D28D9")
    r = 2
    for t in tickets:
        if t["ado"]:
            _write_row(ws8, r, t, with_reason=False); r += 1

    # -- Aging View
    ws9 = wb.create_sheet("Aging View")
    _set_widths(ws9, W14)
    r = 1
    for bkt in ["30d+","15-30d","8-14d","0-7d"]:
        _write_header(ws9, r, H14, "475569"); r += 1
        for t in tickets:
            if t["bucket"] == bkt:
                _write_row(ws9, r, t, with_reason=True); r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_html(tickets: list[dict], today: date) -> str:
    """Return full HTML report (table-based inline styles, email-compatible)."""
    from collections import defaultdict

    mo_long = ["January","February","March","April","May","June",
               "July","August","September","October","November","December"]
    title_date = f"{mo_long[today.month-1]} {today.day}, {today.year}"

    by_status: dict[str, list] = {}
    for t in tickets:
        by_status.setdefault(t["status"], []).append(t)

    open_ = by_status.get("Open", [])
    ip    = by_status.get("In Progress", [])
    oh    = by_status.get("On Hold", [])
    arc   = by_status.get("Awaiting Resolution Confirmation", [])

    n_new = len(open_); n_ip = len(ip); n_oh = len(oh); n_arc = len(arc)

    # L2: only Platinum/Gold/Silver with ADO
    n_l2 = sum(1 for t in tickets
               if t["ado"] and t["band"] in ("Platinum","Gold","Silver"))

    ag = {"30d+":0,"15-30d":0,"8-14d":0,"0-7d":0}
    for t in tickets:
        ag[t["bucket"]] += 1

    n_plat   = sum(1 for t in tickets if t["band"] == "Platinum")
    n_gold   = sum(1 for t in tickets if t["band"] == "Gold")
    n_silver = sum(1 for t in tickets if t["band"] == "Silver")

    # ── Pill helpers ──────────────────────────────────────────────────────────
    _AGE_PILL = {
        "0-7d":  "background:#F0FDF4;color:#14532D",
        "8-14d": "background:#FEFCE8;color:#713F12",
        "15-30d":"background:#FEF3C7;color:#B45309",
        "30d+":  "background:#FEE2E2;color:#B91C1C",
    }
    _PRI_PILL = {
        "P2": "background:#FFF7ED;color:#C2410C",
        "P3": "background:#F1F5F9;color:#475569",
    }
    _STAT_PILL = {
        "Open":                            "background:#EFF6FF;color:#1D4ED8",
        "In Progress":                     "background:#F0FDF4;color:#15803D",
        "On Hold":                         "background:#FFFBEB;color:#B45309",
        "Awaiting Resolution Confirmation":"background:#EDE9FE;color:#6D28D9",
    }
    _STAT_BADGE_STYLE = {
        "Open":                            "background:#EFF6FF;color:#1D4ED8",
        "In Progress":                     "background:#F0FDF4;color:#15803D",
        "On Hold":                         "background:#FFFBEB;color:#B45309",
        "Awaiting Resolution Confirmation":"background:#EDE9FE;color:#6D28D9",
    }
    _ROW_BG = {
        ("On Hold","30d+"):      "#FFF0F0",
        ("On Hold","15-30d"):    "#FFFBF0",
        ("On Hold","8-14d"):     "#FFFBF0",
        ("In Progress","8-14d"): "#FEFFF0",
    }
    _BAND_BADGE_HTML = {
        "Platinum": '<span style="font-size:9px;font-weight:700;background:#1E293B;color:#CBD5E1;border:1px solid #334155;border-radius:4px;padding:2px 7px;">&#x1F451; Platinum</span>',
        "Gold":     '<span style="font-size:9px;font-weight:700;background:#FFFBEB;color:#92400E;border:1px solid #FDE68A;border-radius:4px;padding:2px 7px;">&#x2B50; Gold</span>',
        "Silver":   '<span style="font-size:9px;font-weight:700;background:#F1F5F9;color:#475569;border:1px solid #CBD5E1;border-radius:4px;padding:2px 7px;">&#x1F948; Silver</span>',
        "Bronze":   '<span style="font-size:9px;font-weight:700;background:#FFF7ED;color:#9A3412;border:1px solid #FED7AA;border-radius:4px;padding:2px 7px;">&#x1F536; Bronze</span>',
    }

    _TH = (
        '<tr style="background:#F8FAFC;">'
        '<th style="padding:5px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;text-align:left;border-bottom:1px solid #E4E8EF;width:7%;">Ticket</th>'
        '<th style="padding:5px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;text-align:left;border-bottom:1px solid #E4E8EF;width:22%;">Subject</th>'
        '<th style="padding:5px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;text-align:left;border-bottom:1px solid #E4E8EF;width:8%;">Priority</th>'
        '<th style="padding:5px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;text-align:left;border-bottom:1px solid #E4E8EF;width:13%;">Status</th>'
        '<th style="padding:5px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;text-align:left;border-bottom:1px solid #E4E8EF;width:5%;">Age</th>'
        '<th style="padding:5px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;text-align:left;border-bottom:1px solid #E4E8EF;width:5%;">Team</th>'
        '<th style="padding:5px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;text-align:left;border-bottom:1px solid #E4E8EF;width:6%;">ADO #</th>'
        '<th style="padding:5px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;text-align:left;border-bottom:1px solid #E4E8EF;width:13%;">Raised By</th>'
        '<th style="padding:5px 10px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;text-align:left;border-bottom:1px solid #E4E8EF;width:13%;">Last Update</th>'
        '</tr>'
    )

    def _ticket_rows(lst: list[dict]) -> str:
        rows = []
        for t in lst:
            rbg = _ROW_BG.get((t["status"], t["bucket"]), "#FFFFFF")
            age_sty  = _AGE_PILL[t["bucket"]]
            pri_sty  = _PRI_PILL.get(t["priority"], "background:#F1F5F9;color:#475569")
            stat_sty = _STAT_PILL.get(t["status"], "background:#F1F5F9;color:#475569")
            team_td  = (
                '<span style="font-size:9px;font-weight:700;background:#EDE9FE;color:#6D28D9;padding:2px 5px;border-radius:3px;">L2</span>'
                if t["team"] == "L2" else
                '<span style="font-size:9px;font-weight:600;background:#F1F5F9;color:#64748B;padding:2px 5px;border-radius:3px;">L1</span>'
            )
            ado_td = (
                f'<span style="font-size:10px;font-weight:700;background:#F5F3FF;color:#6D28D9;padding:2px 5px;border-radius:4px;">{t["ado"]}</span>'
                if t["ado"] else ""
            )
            rows.append(
                f'<tr style="background:{rbg};">'
                f'<td style="padding:5px 10px;font-size:11px;font-weight:600;color:#2563EB;border-bottom:1px solid #F1F5F9;">#{t["num"]}</td>'
                f'<td style="padding:5px 10px;font-size:11px;color:#334155;border-bottom:1px solid #F1F5F9;">{t["subject"]}</td>'
                f'<td style="padding:5px 10px;border-bottom:1px solid #F1F5F9;"><span style="font-size:10px;font-weight:700;{pri_sty};padding:2px 8px;border-radius:4px;">{t["priority"]}</span></td>'
                f'<td style="padding:5px 10px;border-bottom:1px solid #F1F5F9;"><span style="font-size:10px;font-weight:600;{stat_sty};padding:2px 7px;border-radius:10px;white-space:nowrap;">{t["status"]}</span></td>'
                f'<td style="padding:5px 10px;border-bottom:1px solid #F1F5F9;"><span style="font-size:10px;font-weight:600;{age_sty};padding:2px 6px;border-radius:10px;">{t["age"]}d</span></td>'
                f'<td style="padding:5px 10px;border-bottom:1px solid #F1F5F9;">{team_td}</td>'
                f'<td style="padding:5px 10px;border-bottom:1px solid #F1F5F9;">{ado_td}</td>'
                f'<td style="padding:5px 10px;font-size:10px;color:#475569;border-bottom:1px solid #F1F5F9;">{t["contact"]}</td>'
                f'<td style="padding:5px 10px;font-size:10px;color:#64748B;border-bottom:1px solid #F1F5F9;">{t["last_updated"]}</td>'
                f'</tr>'
            )
        return "".join(rows)

    def _status_counts(lst: list[dict]) -> str:
        counts: dict[str, int] = defaultdict(int)
        for t in lst:
            counts[t["status"]] += 1
        parts = []
        for s in ["Awaiting Resolution Confirmation","On Hold","In Progress","Open"]:
            if counts[s]:
                sty = _STAT_BADGE_STYLE[s]
                parts.append(f'<span style="font-size:10px;{sty};padding:2px 6px;border-radius:4px;margin-left:4px;">{counts[s]} {s}</span>')
        return "".join(parts)

    def _account_block(name: str, band: str, lst: list[dict]) -> str:
        badge = _BAND_BADGE_HTML.get(band, _BAND_BADGE_HTML["Bronze"])
        sc    = _status_counts(lst)
        rows  = _ticket_rows(lst)
        return (
            '<tr><td style="padding-bottom:14px;">'
            '<table width="100%" cellpadding="0" cellspacing="0" border="0" '
            'style="background:#fff;border:1px solid #E4E8EF;border-radius:10px;margin-bottom:10px;">'
            '<tr><td style="background:#F8FAFC;border-bottom:1px solid #E4E8EF;padding:8px 14px;">'
            '<table width="100%" cellpadding="0" cellspacing="0" border="0"><tr>'
            f'<td style="font-size:13px;font-weight:700;color:#0F172A;">{name}&nbsp;{badge}</td>'
            f'<td align="right">{sc}</td>'
            '</tr></table></td></tr>'
            '<tr><td><table width="100%" cellpadding="0" cellspacing="0" border="0">'
            f'{_TH}{rows}'
            '</table></td></tr></table></td></tr>'
        )

    # HTML shows Platinum / Gold / Silver (Bronze goes to Excel only)
    plat_grp:   dict[str, list] = defaultdict(list)
    gold_grp:   dict[str, list] = defaultdict(list)
    silver_grp: dict[str, list] = defaultdict(list)

    for t in tickets:
        if t["band"] == "Platinum":
            plat_grp[t["display"]].append(t)
        elif t["band"] == "Gold":
            gold_grp[t["display"]].append(t)
        elif t["band"] == "Silver":
            silver_grp[t["display"]].append(t)
        # Bronze / Unknown excluded from HTML

    sections = ""
    for name, lst in sorted(plat_grp.items()):
        sections += _account_block(name, "Platinum", lst)
    for name, lst in sorted(gold_grp.items()):
        sections += _account_block(name, "Gold", lst)
    for name, lst in sorted(silver_grp.items()):
        sections += _account_block(name, "Silver", lst)

    # ── Assemble full HTML ────────────────────────────────────────────────────
    return (
        '<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">'
        f'<title>Daily Incident Report · {title_date}</title></head>\n'
        '<body style="margin:0;padding:0;background:#F4F6F9;font-family:Arial,sans-serif;font-size:13px;color:#1A2035;">\n'
        '<table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#F4F6F9;">'
        '<tr><td align="center" style="padding:20px 16px 40px;">\n'
        '<table width="980" cellpadding="0" cellspacing="0" border="0" style="max-width:980px;width:100%;">\n'

        # Header
        '<tr><td style="background:#fff;border:1px solid #E4E8EF;border-radius:10px;padding:18px 28px 16px;">'
        '<div style="font-size:17px;font-weight:700;color:#0F172A;">Daily Incident Report</div>'
        f'<div style="font-size:11px;color:#64748B;margin-top:3px;"><b style="color:#334155;">Period:</b>&nbsp;{title_date} · 18:00 IST&nbsp;&nbsp;<b style="color:#334155;">Dept:</b>&nbsp;CoreStack Support</div>'
        '</td></tr><tr><td height="14"></td></tr>\n'

        # Summary label
        '<tr><td style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:#94A3B8;padding-bottom:10px;">Summary</td></tr>\n'

        # 5 status cards
        '<tr><td style="padding-bottom:14px;"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr>'
        f'<td width="20%" style="padding-right:10px;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#fff;border:1px solid #E4E8EF;border-radius:10px;border-top:3px solid #3B82F6;"><tr><td style="padding:12px 14px;"><div style="font-size:10px;font-weight:600;color:#64748B;text-transform:uppercase;">New</div><div style="font-size:26px;font-weight:700;color:#3B82F6;line-height:1.1;margin:4px 0 2px;">{n_new}</div><div style="font-size:10px;color:#94A3B8;">opened</div></td></tr></table></td>'
        f'<td width="20%" style="padding-right:10px;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#fff;border:1px solid #E4E8EF;border-radius:10px;border-top:3px solid #10B981;"><tr><td style="padding:12px 14px;"><div style="font-size:10px;font-weight:600;color:#64748B;text-transform:uppercase;">In Progress</div><div style="font-size:26px;font-weight:700;color:#10B981;line-height:1.1;margin:4px 0 2px;">{n_ip}</div><div style="font-size:10px;color:#94A3B8;">being worked</div></td></tr></table></td>'
        f'<td width="20%" style="padding-right:10px;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#fff;border:1px solid #E4E8EF;border-radius:10px;border-top:3px solid #F59E0B;"><tr><td style="padding:12px 14px;"><div style="font-size:10px;font-weight:600;color:#64748B;text-transform:uppercase;">On Hold</div><div style="font-size:26px;font-weight:700;color:#F59E0B;line-height:1.1;margin:4px 0 2px;">{n_oh}</div><div style="font-size:10px;color:#94A3B8;">pending / monitoring</div></td></tr></table></td>'
        f'<td width="20%" style="padding-right:10px;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#fff;border:1px solid #E4E8EF;border-radius:10px;border-top:3px solid #8B5CF6;"><tr><td style="padding:12px 14px;"><div style="font-size:10px;font-weight:600;color:#64748B;text-transform:uppercase;">Awaiting Confirmation</div><div style="font-size:26px;font-weight:700;color:#8B5CF6;line-height:1.1;margin:4px 0 2px;">{n_arc}</div><div style="font-size:10px;color:#94A3B8;">awaiting customer</div></td></tr></table></td>'
        f'<td width="20%" style="padding-right:0px;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#fff;border:1px solid #E4E8EF;border-radius:10px;border-top:3px solid #EF4444;"><tr><td style="padding:12px 14px;"><div style="font-size:10px;font-weight:600;color:#64748B;text-transform:uppercase;">With L2 (PGS)</div><div style="font-size:26px;font-weight:700;color:#EF4444;line-height:1.1;margin:4px 0 2px;">{n_l2}</div><div style="font-size:10px;color:#94A3B8;">ADO linked</div></td></tr></table></td>'
        '</tr></table></td></tr>\n'

        # Aging label + cards
        '<tr><td style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:#94A3B8;padding-bottom:10px;">Open Ticket Aging</td></tr>\n'
        '<tr><td style="padding-bottom:14px;"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr>'
        f'<td width="25%" style="padding-right:10px;"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#FEF2F2;border:1px solid #FECACA;border-radius:8px;"><tr><td style="padding:10px 14px;"><table width="100%" cellpadding="0" cellspacing="0"><tr><td style="font-size:11px;font-weight:600;color:#991B1B;">30d+</td><td align="right" style="font-size:26px;font-weight:700;color:#DC2626;">{ag["30d+"]}</td></tr></table></td></tr></table></td>'
        f'<td width="25%" style="padding-right:10px;"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#FFFBEB;border:1px solid #FDE68A;border-radius:8px;"><tr><td style="padding:10px 14px;"><table width="100%" cellpadding="0" cellspacing="0"><tr><td style="font-size:11px;font-weight:600;color:#92400E;">15-30d</td><td align="right" style="font-size:26px;font-weight:700;color:#D97706;">{ag["15-30d"]}</td></tr></table></td></tr></table></td>'
        f'<td width="25%" style="padding-right:10px;"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#FEFCE8;border:1px solid #FEF08A;border-radius:8px;"><tr><td style="padding:10px 14px;"><table width="100%" cellpadding="0" cellspacing="0"><tr><td style="font-size:11px;font-weight:600;color:#713F12;">8-14d</td><td align="right" style="font-size:26px;font-weight:700;color:#CA8A04;">{ag["8-14d"]}</td></tr></table></td></tr></table></td>'
        f'<td width="25%"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#F0FDF4;border:1px solid #BBF7D0;border-radius:8px;"><tr><td style="padding:10px 14px;"><table width="100%" cellpadding="0" cellspacing="0"><tr><td style="font-size:11px;font-weight:600;color:#14532D;">0-7d</td><td align="right" style="font-size:26px;font-weight:700;color:#16A34A;">{ag["0-7d"]}</td></tr></table></td></tr></table></td>'
        '</tr></table></td></tr>\n'

        # Band label + cards
        '<tr><td style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:#94A3B8;padding-bottom:10px;">Tickets by Account Band</td></tr>\n'
        '<tr><td style="padding-bottom:14px;"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr>'
        f'<td width="33%" style="padding-right:10px;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#1E293B;border:1px solid #334155;border-radius:10px;"><tr><td style="padding:14px 16px;"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="font-size:11px;font-weight:700;color:#CBD5E1;text-transform:uppercase;">&#x1F451; Platinum</td><td align="right" style="font-size:28px;font-weight:700;color:#CBD5E1;">{n_plat}</td></tr></table></td></tr></table></td>'
        f'<td width="33%" style="padding-right:10px;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#451A03;border:1px solid #78350F;border-radius:10px;"><tr><td style="padding:14px 16px;"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="font-size:11px;font-weight:700;color:#FBBF24;text-transform:uppercase;">&#x2B50; Gold</td><td align="right" style="font-size:28px;font-weight:700;color:#FBBF24;">{n_gold}</td></tr></table></td></tr></table></td>'
        f'<td width="33%" style="padding-right:10px;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#1E3A5F;border:1px solid #1E4976;border-radius:10px;"><tr><td style="padding:14px 16px;"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="font-size:11px;font-weight:700;color:#93C5FD;text-transform:uppercase;">&#x1F948; Silver</td><td align="right" style="font-size:28px;font-weight:700;color:#93C5FD;">{n_silver}</td></tr></table></td></tr></table></td>'
        '</tr></table></td></tr>\n'

        # Section label
        '<tr><td style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:#94A3B8;padding-bottom:10px;">Ticket Detail by Account (Platinum · Gold · Silver)</td></tr>\n'

        # Reference legend
        '<tr><td style="padding-bottom:14px;"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#F8FAFC;border:1px solid #E4E8EF;border-radius:10px;"><tr><td style="padding:14px 18px;">'
        '<div style="font-size:10px;font-weight:700;color:#64748B;text-transform:uppercase;letter-spacing:0.06em;margin-bottom:10px;">Reference</div>'
        '<table width="100%" cellpadding="0" cellspacing="0" border="0"><tr>'
        '<td width="50%" valign="top" style="padding-right:24px;border-right:1px solid #E4E8EF;">'
        '<div style="font-size:10px;font-weight:700;color:#B45309;margin-bottom:4px;">&#x1F7E1; On Hold</div>'
        '<div style="font-size:10px;color:#64748B;line-height:1.7;">Ticket kept for monitoring, backfill in progress, pending information from the customer or third parties like Azure / AWS / GCP / OCI.</div>'
        '</td>'
        '<td width="50%" valign="top" style="padding-left:24px;">'
        '<div style="font-size:10px;font-weight:700;color:#6D28D9;margin-bottom:4px;">&#x1F7E3; Awaiting Resolution Confirmation</div>'
        '<div style="font-size:10px;color:#64748B;line-height:1.7;">Issue fixed / clarification has been communicated. Awaiting resolution confirmation from the customer for closure.</div>'
        '</td>'
        '</tr></table>'
        '</td></tr></table></td></tr>\n'

        # Account sections
        + sections +

        # Excel download link footer
        f'<tr><td><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#F0F9FF;border:1px solid #BAE6FD;border-radius:10px;">'
        f'<tr><td style="padding:14px 20px;"><span style="font-size:11px;color:#64748B;">&#128202; Download Excel version: </span>'
        f'<a href="CS_Daily_Incident_Report_{today.strftime("%Y%m%d")}.xlsx" style="font-size:11px;font-weight:600;color:#2563EB;text-decoration:none;">CS_Daily_Incident_Report_{today.strftime("%Y%m%d")}.xlsx &#8594;</a>'
        f'</td></tr></table></td></tr>\n'

        '</table></td></tr></table></body></html>'
    )
