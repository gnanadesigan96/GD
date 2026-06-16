from datetime import date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

TODAY = date(2026, 6, 16)

# fmt: (num, subject, priority, acct, cust, ado, created, last_mod, contact, assignee, status, reason)

OPEN = [
    ("201987","Sonata - AWS newly created resource report issue","P2","CoreStack_CS","Sonata","","2026-04-22","Jun 16","Deovrat Soman","Avinash Naidu","Open",""),
    ("219258","Firing: High Priority MSProd App Server Memory Utilisation above 90%","P3","Corestack","internal","","2026-06-08","Jun 08","Notify SRE Ops","—","Open",""),
]

IP = [
    ("206833","Mar26 usage for Mitsui Chemicals","P3","Synoptek","synoptek","129985","2026-05-07","Jun 15","Stacey Zborowski","Nithin Ram","In Progress",""),
    ("211192","FinOps Email Notification Issue (Clone)","P3","tatacommunications","tcl","125029","2026-05-19","Jun 09","Sunilkumar S","PremKumar B","In Progress",""),
    ("216750","FW: Production Environment IFoundry5X","P3","tatacommunications","tcl","132646","2026-06-02","Jun 09","Sunilkumar S","PremKumar B","In Progress",""),
    ("216808","Difference Between CoreStack Recommendation and Azure Calculator Pricing","P3","Kyndryl","kyndryl","132316","2026-06-02","Jun 04","Randhir Kumar","PremKumar B","In Progress",""),
    ("217515","FW: Cost Mismatch in recommendation","P3","CoreStack","kyndryl","","2026-06-04","Jun 04","Nagalakshmi N","Ganga Reddy","In Progress",""),
    ("219541","Trustech - Finops not triggered","P2","CoreStack_CS","internal","","2026-06-09","Jun 09","Krishna Kumar VJ","Ganga Reddy","In Progress",""),
    ("219898","Deployment Status Confirmation Required","P3","cloud-kinetics","cloudkinetics","","2026-06-10","Jun 15","Service Assurance","PremKumar B","In Progress",""),
    ("220105","Merged cells in cost recommendation Report","P3","CoreStack","kyndryl","133816","2026-06-10","Jun 10","Nagalakshmi N","PremKumar B","In Progress",""),
    ("220546","ODP/Blackstone - Bus Patrol","P3","Corestack","internal","","2026-06-11","Jun 11","Jayven Couch","Nithin Ram","In Progress",""),
    ("220999","Request for Investigation – OCI Cost Processing","P3","core42","internal","133976","2026-06-12","Jun 12","Muthu D","PremKumar B","In Progress",""),
    ("222003","SHI Locuz - Need Assistance in Compliance Execution & Scheduling","P3","CoreStack","internal","","2026-06-15","Jun 15","Nagalakshmi N","Deepesh H","In Progress",""),
    ("222040","Billing Amount Difference between GCP and Core Stock","P3","ltts","internal","","2026-06-15","Jun 15","Kaustubh M","Deepesh H","In Progress",""),
    ("222173","PoV Exide : Unable to execute CIS compliance standard and executed standards has many errors","P3","CoreStack","internal","","2026-06-15","Jun 15","Nagalakshmi N","Deepesh H","In Progress",""),
    ("222218","Enable FinOps Governance Summary Report | QapiPlus","P3","aliando","internal","","2026-06-15","Jun 15","Ajit Thapa","Nithin Ram","In Progress",""),
    ("222257","Arcera - replace payg to Csp onboarding","P3","CoreStack_CS","internal","","2026-06-15","Jun 15","Krishna Kumar VJ","Nithin Ram","In Progress",""),
    ("222287","Confirmation Request – AI (LLM) Cost Details","P3","otsuka-us","Otsuka","","2026-06-15","Jun 15","Rajkumar Uppu","Nithin Ram","In Progress",""),
    ("222461","Kyndryl SaaS Okta SSO Configuration","P3","CoreStack","kyndryl","","2026-06-16","Jun 16","Nagalakshmi N","Avinash Naidu","In Progress",""),
    ("222463","Mismatch in Cost Figures Between Azure Cost Management and CoreStack FinOps Dashboard","P3","sonata-software","Sonata","","2026-06-16","Jun 16","Thadi Swathi","Avinash Naidu","In Progress",""),
]

OH = [
    ("205888","Need assistance to update creds for EA account for Trinity College","P2","au.logicalis","logicallis","132186","2026-05-05","Jun 12","Kamran Wahid","PremKumar B","On Hold","Customer is checking access permissions with their admin."),
    ("211893","Re: Core stock Finops Dashboard cost differ","P2","neurealm","","","2026-05-21","Jun 15","Parthasarathy K","Avinash Naidu","On Hold","Customer needs to raise the support case with Azure. This is an issue from Azure side."),
    ("211895","Re: Core stock Finops Dashboard cost differ for GCP","P3","neurealm","internal","131068","2026-05-21","Jun 15","Parthasarathy K","Nithin Ram","On Hold","Steps provided customer has to implement the changes."),
    ("211954","AWS Accounts transfer from INH to ISO Tenant","P3","otsuka-us","Otsuka","","2026-05-21","May 22","Rajkumar Uppu","Aadhithya S","On Hold","Awaiting confirmation from Ashok to proceed with backfilling of the cost data for these 3 accounts."),
    ("217785","RDS Snapshot Not Created on May 31","P3","cloud-kinetics","cloudkinetics","132850","2026-06-05","Jun 11","Service Assurance","Gnanadesigan A","On Hold","We have stated that we do not have sufficient logs to troubleshoot further and the customer is checking internally."),
    ("217961","Sonata - CSP accounts not showing up","P3","CoreStack_CS","Sonata","132668","2026-06-05","Jun 09","Deovrat Soman","PremKumar B","On Hold","The initial reported issue has been resolved, however while loading the dashboard we are encountering errors. Engineering team suspects the issue is due to missing currency."),
    ("217990","Cloud.corestack.io is slow across all pages","P3","CoreStack","internal","","2026-06-05","Jun 12","Satyabrat","Ganga Reddy","On Hold","We are awaiting response from Pendo team."),
    ("219147","RE: RE:[CASE] CUR Backfill","P3","sonata-software","Sonata","","2026-06-08","Jun 13","Raghavan P","Nithin Ram","On Hold","Flow currently being tested and 1 account works as expected. Will proceed with the remaining."),
    ("219360","cloud.corestack.io","P3","CoreStack","internal","","2026-06-08","Jun 09","Satyabrat","Nithin Ram","On Hold","NA"),
]

ARC = [
    ("214686","No cost data for TreeRing (AEMCS)","P3","Corestack","internal","","2026-05-27","Jun 11","Jayven Couch","Logesh S","Awaiting Resolution Confirmation","Waiting for the ticket owner to confirm."),
    ("216586","Unable to onboard Snowflake in CS4CS","P3","CoreStack","internal","","2026-06-01","Jun 05","Anaranya Bagchi","Ganga Reddy","Awaiting Resolution Confirmation","Waiting for Anaranya's availability for call."),
    ("217606","Login Issue with CoreStack Tool","P3","otsuka-us","Otsuka","","2026-06-04","Jun 12","Rajkumar Uppu","Ganga Reddy","Awaiting Resolution Confirmation","Customer pinged in teams and asked to hold for 1 day."),
    ("219361","Filtrona Finops Dashboard Unallocated Resource Groups","P3","Getronics","Getronics","","2026-06-08","Jun 08","Shashank N","Nithin Ram","Awaiting Resolution Confirmation","NA"),
    ("219377","ODP Corporation MCA Billing Account cost processing errors","P3","Corestack","internal","","2026-06-08","Jun 08","Jayven Couch","Nithin Ram","Awaiting Resolution Confirmation","Awaiting credential refresh to validate the cost process."),
    ("219658","US Prod - Dashboard Not Loading","P3","CoreStack_CS","internal","","2026-06-09","Jun 09","Ashok Kumar E","Nithin Ram","Awaiting Resolution Confirmation","NA"),
    ("219989","Re: Corestack","P3","CoreStack_CS","Sonata","","2026-06-10","Jun 10","Deovrat Soman","PremKumar B","Awaiting Resolution Confirmation","Unable to reproduce the issue, informed the same to the customer and we are awaiting their response."),
    ("220165","Trustedtech - HMH - cost for April","P3","CoreStack_CS","internal","","2026-06-10","Jun 10","Krishna Kumar VJ","Nithin Ram","Awaiting Resolution Confirmation","N/A"),
    ("220416","Sonata - Tata Tele CSP processing issue","P2","CoreStack_CS","Sonata","133815","2026-06-11","Jun 12","Deovrat Soman","PremKumar B","Awaiting Resolution Confirmation","Cost has been processed and we are now awaiting customer's confirmation."),
    ("220491","FW: Resources Cost - Beside Tagged and untagged - LifeLabs","P3","CoreStack","kyndryl","133656","2026-06-11","Jun 15","Nagalakshmi N","Avinash Naidu","Awaiting Resolution Confirmation","NA"),
    ("220875","Dashboard slowness","P2","Synopsys","synopsys","133875","2026-06-12","Jun 15","Ranjitha Thota","Ganga Reddy","Awaiting Resolution Confirmation","Resolved and awaiting customer confirmation."),
    ("221054","US SaaS - Kyndryl - Default dashboards not visible for tenant/finops_admin","P3","CoreStack","kyndryl","","2026-06-12","Jun 12","Nagalakshmi N","Nithin Ram","Awaiting Resolution Confirmation","NA"),
    ("222219","Request for Azure DevOps Cost Details at Resource Level","P3","otsuka-us","Otsuka","","2026-06-15","Jun 15","Rajkumar Uppu","Nithin Ram","Awaiting Resolution Confirmation","Awaiting customer confirmation on findings."),
]

ALL_TICKETS = OPEN + IP + OH + ARC

BAND_MAP_ACCT = {
    "neurealm":"Platinum","otsuka-us":"Gold","Kyndryl":"Gold","tatacommunications":"Gold",
    "au.logicalis":"Gold","Synopsys":"Gold","Synoptek":"Gold","Getronics":"Silver",
    "cloud-kinetics":"Silver",
}
CUST_GOLD   = {"kyndryl","Otsuka","synopsys","synoptek","tcl","logicallis"}
CUST_SILVER = {"Getronics","cloudkinetics"}

def get_band(acct, cust):
    k = acct.lower()
    if k == "neurealm": return "Platinum"
    if acct in BAND_MAP_ACCT: return BAND_MAP_ACCT[acct]
    if cust in CUST_GOLD: return "Gold"
    if cust in CUST_SILVER: return "Silver"
    return "Bronze"

def get_disp(acct, cust):
    m = {
        "neurealm":"Neurealm","otsuka-us":"Otsuka","Kyndryl":"Kyndryl",
        "tatacommunications":"Tata Communications","au.logicalis":"Logicalis",
        "Synopsys":"Synopsys","Synoptek":"Synoptek","Getronics":"Getronics",
        "cloud-kinetics":"Cloud Kinetics","sonata-software":"Sonata","ltts":"LTTS",
        "gevernova":"GE Vernova","core42":"Core42","aliando":"Aliando",
    }
    if acct in m: return m[acct]
    cm = {
        "kyndryl":"Kyndryl","Otsuka":"Otsuka","synopsys":"Synopsys","synoptek":"Synoptek",
        "tcl":"Tata Communications","logicallis":"Logicalis","Getronics":"Getronics",
        "cloudkinetics":"Cloud Kinetics","Sonata":"Sonata","LTTS":"LTTS",
        "GE Vernova":"GE Vernova","internal":"CoreStack (Internal)",
    }
    return cm.get(cust, acct)

def age(d): return (TODAY - date.fromisoformat(d)).days
def bucket(a):
    if a<=7:  return "0-7d"
    if a<=14: return "8-14d"
    if a<=30: return "15-30d"
    return "30d+"

def created_disp(d):
    mo = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    dt = date.fromisoformat(d)
    return f"{mo[dt.month-1]} {dt.day:02d}"

# ── Fills & Fonts ──────────────────────────────────────────────────────────────
def f(hex6): return PatternFill("solid", fgColor=hex6)
def font(hex6, bold=False, sz=10): return Font(color=hex6, bold=bold, size=sz)

ROW_BG = {
    ("On Hold","30d+"):      "FFF0F0",
    ("On Hold","15-30d"):    "FFFBF0",
    ("On Hold","8-14d"):     "FFFBF0",
    ("In Progress","8-14d"): "FEFFF0",
}
def row_bg(status, bkt): return ROW_BG.get((status, bkt), "FFFFFF")

AGE_STYLES    = {"0-7d":("F0FDF4","14532D"),"8-14d":("FEFCE8","713F12"),
                 "15-30d":("FEF3C7","B45309"),"30d+":("FEE2E2","B91C1C")}
PRI_STYLES    = {"P2":("FFF7ED","C2410C"),"P3":("F1F5F9","475569")}
STATUS_STYLES = {"Open":("DBEAFE","1D4ED8"),"In Progress":("DCFCE7","15803D"),
                 "On Hold":("FEF9C3","B45309"),
                 "Awaiting Resolution Confirmation":("EDE9FE","6D28D9")}
BAND_STYLES   = {"Platinum":("CBD5E1","1E293B"),"Gold":("FEF9C3","92400E"),
                 "Silver":("DBEAFE","1E4976"),"Bronze":("FFF7ED","9A3412")}
TEAM_STYLES   = {"L2":("EDE9FE","6D28D9"),"L1":("F1F5F9","64748B")}

al_center = Alignment(horizontal="center",vertical="center",wrap_text=True)
al_left   = Alignment(horizontal="left",  vertical="center",wrap_text=True)

def write_header(ws, row, headers, fill_hex):
    for i,h in enumerate(headers,1):
        c = ws.cell(row,i,h)
        c.fill=f(fill_hex); c.font=Font(color="FFFFFF",bold=True,size=10)
        c.alignment=al_center
    ws.row_dimensions[row].height=25

def write_row(ws, row_num, t, with_reason=True):
    num,subj,pri,acct,cust,ado,created,mod,contact,assignee,status,reason = t
    a = age(created); bkt = bucket(a)
    band = get_band(acct,cust)
    # Blackstone override
    if "blackstone" in subj.lower() or "blackstone" in acct.lower():
        band = "Bronze"
        disp = "Blackstone"
    else:
        disp = get_disp(acct,cust)
    team = "L2" if ado else "L1"
    rbg  = row_bg(status,bkt)

    af,afont = AGE_STYLES[bkt]
    pf,pfont = PRI_STYLES.get(pri,("F1F5F9","475569"))
    sf,sfont = STATUS_STYLES.get(status,("FFFFFF","000000"))
    bf,bfont = BAND_STYLES.get(band,("FFF7ED","9A3412"))

    cells = [
        (f"#{num}",   rbg,     "2563EB", True),
        (disp,        rbg,     "0F172A", False),
        (band,        bf,      bfont,    True),
        (subj,        rbg,     "0F172A", False),
        (pri,         pf,      pfont,    True),
        (status,      sf,      sfont,    True),
        (f"{a}d",     af,      afont,    True),
        (bkt,         af,      afont,    False),
        (team,        "EDE9FE" if team=="L2" else "F1F5F9",
                      "6D28D9" if team=="L2" else "64748B", team=="L2"),
        (ado,         "F5F3FF" if ado else "FFFFFF", "6D28D9", bool(ado)),
        (contact,     "F8FAFC","334155",False),
    ]
    if with_reason:
        cells.append((reason,  "FFFBEB","92400E",False))
        cells.append((mod,     rbg,    "0F172A",False))
    else:
        cells.append((mod,     rbg,    "0F172A",False))
    cells.append((created_disp(created), rbg, "0F172A", False))

    for col_idx,(val,fill_hex,font_hex,bold) in enumerate(cells,1):
        c = ws.cell(row_num,col_idx,val)
        c.fill=f(fill_hex); c.font=Font(color=font_hex,bold=bold,size=9)
        c.alignment=al_left
    ws.row_dimensions[row_num].height=30

def set_widths(ws, widths):
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w

W14 = [10,18,10,40,8,26,8,10,6,10,16,36,16,12]
W13 = [10,18,10,44,8,26,8,10,6,10,16,16,12]
H14 = ["Ticket #","Customer","Band","Subject","Priority","Status","Age","Bucket","Team","ADO #","Raised By","Reason","Last Updated","Created"]
H13 = ["Ticket #","Customer","Band","Subject","Priority","Status","Age","Bucket","Team","ADO #","Raised By","Last Updated","Created"]

wb = Workbook()

# ── Summary ───────────────────────────────────────────────────────────────────
ws1 = wb.active; ws1.title="Summary"
ws1.column_dimensions["A"].width=32
ws1.column_dimensions["B"].width=10
ws1.column_dimensions["C"].width=12

ws1["A1"]="Daily Incident Report — June 16, 2026"
ws1["A1"].font=Font(color="0F172A",bold=True,size=12)
ws1["A2"]="Period: June 16, 2026 · 18:00 IST"
ws1["A2"].font=Font(color="64748B",size=10)
ws1.row_dimensions[1].height=22; ws1.row_dimensions[2].height=18
ws1.row_dimensions[3].height=8

write_header(ws1,4,["Status","Count","% of Total"],"1E293B")
total=len(ALL_TICKETS)
n_open=len(OPEN); n_ip=len(IP); n_oh=len(OH); n_arc=len(ARC)
for i,(s,n) in enumerate([("Open",n_open),("In Progress",n_ip),("On Hold",n_oh),
                            ("Awaiting Resolution Confirmation",n_arc),("Total",total)],5):
    ws1[f"A{i}"]=s; ws1[f"B{i}"]=n
    ws1[f"C{i}"]=f"{n/total*100:.1f}%" if s!="Total" else "100%"
    ws1.row_dimensions[i].height=22
    if s!="Total":
        sf,sfont=STATUS_STYLES.get(s,("FFFFFF","000000"))
        for col in ["A","B","C"]:
            c=ws1[f"{col}{i}"]; c.fill=f(sf); c.font=Font(color=sfont,bold=True,size=10)
    else:
        for col in ["A","B","C"]: ws1[f"{col}{i}"].font=Font(bold=True,size=10)

# ── All Tickets ───────────────────────────────────────────────────────────────
ws2=wb.create_sheet("All Tickets"); set_widths(ws2,W14)
write_header(ws2,1,H14,"1E293B")
for r,t in enumerate(ALL_TICKETS,2): write_row(ws2,r,t,with_reason=True)

# ── Platinum Gold Silver ──────────────────────────────────────────────────────
ws3=wb.create_sheet("Platinum Gold Silver"); set_widths(ws3,W14)
write_header(ws3,1,H14,"334155")
r=2
for t in ALL_TICKETS:
    if get_band(t[3],t[4]) in ("Platinum","Gold","Silver"):
        write_row(ws3,r,t,with_reason=True); r+=1

# ── New ───────────────────────────────────────────────────────────────────────
ws4=wb.create_sheet("New"); set_widths(ws4,W13)
write_header(ws4,1,H13,"1D4ED8")
for r,t in enumerate(OPEN,2): write_row(ws4,r,t,with_reason=False)

# ── In Progress ───────────────────────────────────────────────────────────────
ws5=wb.create_sheet("In Progress"); set_widths(ws5,W13)
write_header(ws5,1,H13,"15803D")
for r,t in enumerate(IP,2): write_row(ws5,r,t,with_reason=False)

# ── On Hold ───────────────────────────────────────────────────────────────────
ws6=wb.create_sheet("On Hold"); set_widths(ws6,W14)
write_header(ws6,1,H14,"B45309")
for r,t in enumerate(OH,2): write_row(ws6,r,t,with_reason=True)

# ── Awaiting Resolution ───────────────────────────────────────────────────────
ws7=wb.create_sheet("Awaiting Resolution"); set_widths(ws7,W14)
write_header(ws7,1,H14,"6D28D9")
for r,t in enumerate(ARC,2): write_row(ws7,r,t,with_reason=True)

# ── L2 Tickets ────────────────────────────────────────────────────────────────
ws8=wb.create_sheet("L2 Tickets"); set_widths(ws8,W13)
write_header(ws8,1,H13,"6D28D9")
r=2
for t in ALL_TICKETS:
    if t[5]: write_row(ws8,r,t,with_reason=False); r+=1

# ── Aging View ────────────────────────────────────────────────────────────────
ws9=wb.create_sheet("Aging View"); set_widths(ws9,W14)
r=1
for bkt in ["30d+","15-30d","8-14d","0-7d"]:
    write_header(ws9,r,H14,"475569"); r+=1
    for t in ALL_TICKETS:
        a=age(t[6])
        if bucket(a)==bkt:
            write_row(ws9,r,t,with_reason=True); r+=1

wb.save("/home/user/GD/CS_Daily_Incident_Report_20260616.xlsx")
print("Excel done ✓")

# ── HTML ──────────────────────────────────────────────────────────────────────
import textwrap

def age_cls(bkt):
    return {"0-7d":"age-0","8-14d":"age-1","15-30d":"age-2","30d+":"age-3"}[bkt]

def stat_cls(status):
    return {"Open":"stat-open","In Progress":"stat-ip","On Hold":"stat-oh",
            "Awaiting Resolution Confirmation":"stat-arc"}.get(status,"")

def stat_label(status):
    return "ARC" if status=="Awaiting Resolution Confirmation" else status

def pri_cls(pri): return "pri-p2" if pri=="P2" else "pri-p3"

def badge_cls(band):
    return {"Platinum":"badge-plat","Gold":"badge-gold","Silver":"badge-silver"}.get(band,"badge-bronze")

def ticket_table(tickets, show_reason=True):
    reason_th = "<th>Reason / Last Update</th>" if show_reason else ""
    rows = []
    for t in tickets:
        num,subj,pri,acct,cust,ado,created,mod,contact,assignee,status,reason = t
        a=age(created); bkt=bucket(a)
        band=get_band(acct,cust)
        if "blackstone" in subj.lower() or "blackstone" in acct.lower():
            disp="Blackstone"
        else:
            disp=get_disp(acct,cust)
        team="L2" if ado else "L1"
        reason_td = f'<td class="reason">{reason or "—"}</td>' if show_reason else ""
        rows.append(f"""<tr>
<td class="ticket-id">#{num}</td>
<td>{disp}</td>
<td><span class="{age_cls(bkt)}">{a}d</span></td>
<td><span class="{pri_cls(pri)}">{pri}</span></td>
<td><span class="{stat_cls(status)}">{stat_label(status)}</span></td>
<td style="max-width:300px;font-size:11px">{subj}</td>
<td><span class="{"ado" if ado else "team-l1"}">{ado or "—"}</span></td>
<td><span class="{"team-l2" if team=="L2" else "team-l1"}">{team}</span></td>
<td style="font-size:11px;color:#64748B">{contact}</td>
<td style="font-size:11px;color:#334155">{assignee}</td>
{reason_td}
<td style="font-size:10px;color:#94A3B8;white-space:nowrap">{mod}</td>
</tr>""")
    reason_th_str = "<th>Reason / Last Update</th>" if show_reason else ""
    return f"""<table>
<thead><tr>
<th>Ticket #</th><th>Customer</th><th>Age</th><th>Pri</th><th>Status</th>
<th>Subject</th><th>ADO #</th><th>Team</th><th>Raised By</th><th>Assignee</th>
{reason_th_str}<th>Modified</th>
</tr></thead>
<tbody>{''.join(rows)}</tbody>
</table>"""

def account_section(name, band, tickets, show_reason=True):
    bc=badge_cls(band)
    badge_label=band if band in ("Platinum","Gold","Silver") else "Bronze"
    return f"""<div class="section">
<div class="section-header">
  <span class="section-title">{name}</span>
  <span class="badge {bc}">{badge_label}</span>
  <span style="font-size:12px;color:#64748B">({len(tickets)} ticket{"s" if len(tickets)!=1 else ""})</span>
</div>
{ticket_table(tickets, show_reason)}
</div>"""

# Build grouped sections
from collections import defaultdict

def group_by_display(tickets):
    g = defaultdict(list)
    for t in tickets:
        if "blackstone" in t[1].lower() or "blackstone" in t[3].lower():
            g["Blackstone"].append(t)
        else:
            g[get_disp(t[3],t[4])].append(t)
    return dict(g)

plat   = [t for t in ALL_TICKETS if get_band(t[3],t[4])=="Platinum"]
gold   = [t for t in ALL_TICKETS if get_band(t[3],t[4])=="Gold"]
silver = [t for t in ALL_TICKETS if get_band(t[3],t[4])=="Silver"]
blackstone = [t for t in ALL_TICKETS if "blackstone" in t[1].lower()]

sections_html = ""
for acct_name, lst in group_by_display(plat).items():
    sections_html += account_section(acct_name, "Platinum", lst)
for acct_name, lst in group_by_display(gold).items():
    sections_html += account_section(acct_name, "Gold", lst)
for acct_name, lst in group_by_display(silver).items():
    sections_html += account_section(acct_name, "Silver", lst)
if blackstone:
    sections_html += account_section("Blackstone", "Bronze", blackstone, show_reason=False)

# Counts
ag = {"30d+":0,"15-30d":0,"8-14d":0,"0-7d":0}
for t in ALL_TICKETS:
    ag[bucket(age(t[6]))] += 1

n_plat  = sum(1 for t in ALL_TICKETS if get_band(t[3],t[4])=="Platinum")
n_gold  = sum(1 for t in ALL_TICKETS if get_band(t[3],t[4])=="Gold")
n_silver= sum(1 for t in ALL_TICKETS if get_band(t[3],t[4])=="Silver")

visible_l2 = sum(1 for t in ALL_TICKETS
                 if t[5] and get_band(t[3],t[4]) in ("Platinum","Gold","Silver")
                 and "blackstone" not in t[1].lower())
print(f"L2 (PGS visible): {visible_l2}")

css = """
body{font-family:'Segoe UI',Arial,sans-serif;background:#F8FAFC;margin:0;padding:20px;color:#0F172A}
.container{max-width:1200px;margin:0 auto;background:#fff;border-radius:12px;box-shadow:0 4px 24px rgba(0,0,0,.08);overflow:hidden}
.header{background:linear-gradient(135deg,#1E293B 0%,#334155 100%);padding:28px 36px;color:#fff}
.header h1{margin:0;font-size:22px;font-weight:700;letter-spacing:-.3px}
.header p{margin:6px 0 0;font-size:13px;opacity:.7}
.summary-grid{display:grid;grid-template-columns:repeat(5,1fr);gap:12px;padding:20px 28px;background:#F8FAFC;border-bottom:1px solid #E2E8F0}
.card{background:#fff;border-radius:8px;padding:14px 16px;border:1px solid #E2E8F0;text-align:center}
.card .label{font-size:11px;color:#64748B;text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px}
.card .value{font-size:26px;font-weight:700;color:#0F172A}
.card.blue{border-top:3px solid #3B82F6}.card.green{border-top:3px solid #22C55E}
.card.yellow{border-top:3px solid #EAB308}.card.purple{border-top:3px solid #A855F7}
.card.l2{border-top:3px solid #6D28D9}
.age-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;padding:0 28px 16px;background:#F8FAFC;border-bottom:1px solid #E2E8F0}
.age-card{border-radius:8px;padding:12px 16px;text-align:center}
.age-card .label{font-size:11px;text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px}
.age-card .value{font-size:22px;font-weight:700}
.age-red{background:#FEE2E2;color:#B91C1C}.age-orange{background:#FEF3C7;color:#B45309}
.age-yellow{background:#FEFCE8;color:#713F12}.age-green{background:#F0FDF4;color:#14532D}
.band-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;padding:0 28px 20px;background:#F8FAFC;border-bottom:1px solid #E2E8F0}
.band-card{border-radius:8px;padding:12px 16px;text-align:center}
.band-card .label{font-size:11px;text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px}
.band-card .value{font-size:22px;font-weight:700}
.plat{background:#CBD5E1;color:#1E293B}.gold{background:#FEF9C3;color:#92400E}.silver{background:#DBEAFE;color:#1E4976}
.section{padding:20px 28px;border-bottom:1px solid #F1F5F9}
.section-header{display:flex;align-items:center;gap:10px;margin-bottom:14px}
.section-title{font-size:15px;font-weight:700;color:#1E293B}
.badge{padding:3px 10px;border-radius:20px;font-size:11px;font-weight:700}
.badge-plat{background:#CBD5E1;color:#1E293B}.badge-gold{background:#FEF9C3;color:#92400E}
.badge-silver{background:#DBEAFE;color:#1E4976}.badge-bronze{background:#FFF7ED;color:#9A3412}
table{width:100%;border-collapse:collapse;font-size:12px}
th{background:#1E293B;color:#fff;padding:8px 10px;text-align:left;font-size:11px;font-weight:600}
td{padding:7px 10px;border-bottom:1px solid #F1F5F9;vertical-align:top}
.ticket-id{color:#2563EB;font-weight:700;white-space:nowrap}
.pri-p2{background:#FFF7ED;color:#C2410C;padding:2px 7px;border-radius:4px;font-size:10px;font-weight:700}
.pri-p3{background:#F1F5F9;color:#475569;padding:2px 7px;border-radius:4px;font-size:10px;font-weight:700}
.stat-open{background:#DBEAFE;color:#1D4ED8;padding:2px 7px;border-radius:4px;font-size:10px;font-weight:600}
.stat-ip{background:#DCFCE7;color:#15803D;padding:2px 7px;border-radius:4px;font-size:10px;font-weight:600}
.stat-oh{background:#FEF9C3;color:#B45309;padding:2px 7px;border-radius:4px;font-size:10px;font-weight:600}
.stat-arc{background:#EDE9FE;color:#6D28D9;padding:2px 7px;border-radius:4px;font-size:10px;font-weight:600}
.age-0{background:#F0FDF4;color:#14532D;padding:2px 6px;border-radius:4px;font-size:10px;font-weight:700}
.age-1{background:#FEFCE8;color:#713F12;padding:2px 6px;border-radius:4px;font-size:10px;font-weight:700}
.age-2{background:#FEF3C7;color:#B45309;padding:2px 6px;border-radius:4px;font-size:10px;font-weight:700}
.age-3{background:#FEE2E2;color:#B91C1C;padding:2px 6px;border-radius:4px;font-size:10px;font-weight:700}
.ado{background:#F5F3FF;color:#6D28D9;padding:2px 6px;border-radius:4px;font-size:10px;font-weight:700}
.team-l2{background:#EDE9FE;color:#6D28D9;padding:2px 6px;border-radius:4px;font-size:10px;font-weight:700}
.team-l1{background:#F1F5F9;color:#64748B;padding:2px 6px;border-radius:4px;font-size:10px}
.reason{font-size:11px;color:#78350F;max-width:280px}
.footer{padding:16px 28px;background:#F8FAFC;text-align:center;font-size:11px;color:#94A3B8}
"""

html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Daily Incident Report — June 16, 2026</title>
<style>{css}</style>
</head>
<body>
<div class="container">
  <div class="header">
    <h1>Daily Incident Report</h1>
    <p>June 16, 2026 &nbsp;·&nbsp; CoreStack Support Team</p>
  </div>
  <div class="summary-grid">
    <div class="card blue"><div class="label">New</div><div class="value">{n_open}</div></div>
    <div class="card green"><div class="label">In Progress</div><div class="value">{n_ip}</div></div>
    <div class="card yellow"><div class="label">On Hold</div><div class="value">{n_oh}</div></div>
    <div class="card purple"><div class="label">Awaiting Confirmation</div><div class="value">{n_arc}</div></div>
    <div class="card l2"><div class="label">With L2 (PGS)</div><div class="value">{visible_l2}</div></div>
  </div>
  <div class="age-grid">
    <div class="age-card age-red"><div class="label">30d+</div><div class="value">{ag["30d+"]}</div></div>
    <div class="age-card age-orange"><div class="label">15-30d</div><div class="value">{ag["15-30d"]}</div></div>
    <div class="age-card age-yellow"><div class="label">8-14d</div><div class="value">{ag["8-14d"]}</div></div>
    <div class="age-card age-green"><div class="label">0-7d</div><div class="value">{ag["0-7d"]}</div></div>
  </div>
  <div class="band-grid">
    <div class="band-card plat"><div class="label">Platinum</div><div class="value">{n_plat}</div></div>
    <div class="band-card gold"><div class="label">Gold</div><div class="value">{n_gold}</div></div>
    <div class="band-card silver"><div class="label">Silver</div><div class="value">{n_silver}</div></div>
  </div>
  {sections_html}
  <div class="footer">
    Generated automatically by CoreStack Support Bot &nbsp;·&nbsp; June 16, 2026 18:00 IST
  </div>
</div>
</body>
</html>"""

with open("/home/user/GD/CS_Daily_Incident_Report_20260616.html","w") as fh:
    fh.write(html)
print("HTML done ✓")
print(f"Totals: Open={n_open}, IP={n_ip}, OH={n_oh}, ARC={n_arc}, Total={total}")
print(f"Aging: {ag}")
print(f"Bands: Plat={n_plat}, Gold={n_gold}, Silver={n_silver}")
